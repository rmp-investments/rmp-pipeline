"""
Build transparent Scoring Reference sheet with Excel formulas
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Load template
wb = openpyxl.load_workbook(r'C:\Users\carso\OneDrive - UCB-O365\Work\RMP\Screener\RMP Screener_PreLinked_v3.xlsx')

# Delete and recreate Scoring Reference sheet
if 'Scoring Reference' in wb.sheetnames:
    del wb['Scoring Reference']
ws = wb.create_sheet('Scoring Reference')

# Styles
header_font_white = Font(bold=True, size=14, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
bold_font = Font(bold=True)
final_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

# Set column widths
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 8

# Helper for XLOOKUP formulas
def xlookup(field):
    return f'=_xlfn.XLOOKUP("{field}",\'Data Inputs\'!$B:$B,\'Data Inputs\'!$C:$C)'

def add_header(row, text):
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = text
    ws[f'A{row}'].font = header_font_white
    ws[f'A{row}'].fill = header_fill
    return row + 1

def add_column_headers(row, headers):
    for col, val in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=val).font = bold_font
        ws.cell(row=row, column=col).fill = subheader_fill
    return row + 1

# Title
ws.merge_cells('A1:F1')
ws['A1'] = 'STAGE 2 SCORING CALCULATOR - TRANSPARENT VERSION'
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = 'All scores calculated via Excel formulas with full traceability'
ws['A3'] = 'Column C = Excel Formula Result | Column D = Python Calculated (for verification)'

row = 5

# Track final score rows for Stage 2 reference
final_score_rows = {}

#===============================================================================
# SUPPLY-DEMAND DRIVERS (5% weight)
#===============================================================================
row = add_header(row, 'SUPPLY-DEMAND DRIVERS (5% weight)')
row = add_column_headers(row, ['Field', 'Formula / Value', 'Excel Calc', 'Python Calc', 'Scale Value', 'Scale Score'])

sd_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Absorption (12mo)', xlookup('12 Mo Absorption Units')
ws[f'D{row}'] = xlookup('SD: Absorption (12mo)')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Delivered (12mo)', xlookup('12 Mo Delivered Units')
ws[f'D{row}'] = xlookup('SD: Delivered (12mo)')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Under Construction', xlookup('Under Construction Units')
ws[f'D{row}'] = xlookup('SD: Under Construction')
row += 1

ws[f'A{row}'] = 'Absorption Ratio'
ws[f'B{row}'] = f'=IF(B{sd_start+1}=0,"N/A",B{sd_start}/B{sd_start+1})'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},2),B{row})'
ws[f'D{row}'] = xlookup('SD: Absorption Ratio')
ws[f'E{row}'], ws[f'F{row}'] = '>=2.0', 10
row += 1

ws[f'A{row}'] = 'Pipeline Ratio'
ws[f'B{row}'] = f'=IF(B{sd_start}<=0,"N/A",B{sd_start+2}/B{sd_start})'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},2),B{row})'
ws[f'D{row}'] = xlookup('SD: Pipeline Ratio')
ws[f'E{row}'], ws[f'F{row}'] = '1.5-2.0', 8
row += 1

ws[f'A{row}'] = 'Base Score'
ws[f'B{row}'] = f'=IF(B{sd_start}<0,3,IF(B{sd_start+1}=0,IF(B{sd_start}>0,9,5),IF(C{sd_start+3}>=2,10,IF(C{sd_start+3}>=1.5,8,IF(C{sd_start+3}>=1,6,IF(C{sd_start+3}>=0.5,4,2))))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('SD: Base Score')
ws[f'E{row}'], ws[f'F{row}'] = '1.0-1.5', 6
row += 1

ws[f'A{row}'] = 'Pipeline Adjustment'
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(C{sd_start+4})),0,IF(C{sd_start+4}>1.5,-2,IF(C{sd_start+4}>1,-1,IF(C{sd_start+4}<0.5,1,0))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('SD: Pipeline Adjustment')
ws[f'E{row}'], ws[f'F{row}'] = '0.5-1.0', 4
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=MAX(1,MIN(10,C{row-2}+C{row-1}))'
ws[f'C{row}'] = f'=B{row}'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = xlookup('SD: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '<0.5', 2
final_score_rows['SD'] = row
row += 2

#===============================================================================
# SUBMARKET SUPPLY-DEMAND OUTLOOK (10% weight)
#===============================================================================
row = add_header(row, 'SUBMARKET SUPPLY-DEMAND OUTLOOK (10% weight)')
row = add_column_headers(row, ['Field', 'Formula / Value', 'Excel Calc', 'Python Calc', 'Adjustment', 'Value'])

so_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Current Vacancy %', xlookup('Submarket Vacancy Rate %')
ws[f'D{row}'] = xlookup('SO: Current Vacancy %')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Vacancy YoY Change', xlookup('Vacancy YoY Change %')
ws[f'D{row}'] = xlookup('SO: Vacancy YoY Change')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Historical Avg Vacancy', xlookup('Vacancy Historical Avg %')
ws[f'D{row}'] = xlookup('SO: Historical Avg Vacancy')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Under Construction', xlookup('Under Construction Units')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Absorption (12mo)', xlookup('12 Mo Absorption Units')
row += 1

ws[f'A{row}'] = 'Vacancy vs Historical'
ws[f'B{row}'] = f'=IF(OR(B{so_start}="",B{so_start+2}=""),"N/A",B{so_start}-B{so_start+2})'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},1),B{row})'
ws[f'D{row}'] = xlookup('SO: Vacancy vs Historical')
ws[f'E{row}'], ws[f'F{row}'] = '>2ppt below hist', '+2'
row += 1

ws[f'A{row}'] = 'Level Adjustment'
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(C{row-1})),0,IF(C{row-1}<=-2,2,IF(C{row-1}<=-0.5,1,IF(C{row-1}<=0.5,0,IF(C{row-1}<=2,-1,-2)))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('SO: Level Adjustment')
ws[f'E{row}'], ws[f'F{row}'] = '0.5-2ppt below', '+1'
row += 1

ws[f'A{row}'] = 'Trend Adjustment (YoY)'
ws[f'B{row}'] = f'=IF(B{so_start+1}="",0,IF(B{so_start+1}<=-1,3,IF(B{so_start+1}<=-0.5,2,IF(B{so_start+1}<0,1,IF(B{so_start+1}=0,0,IF(B{so_start+1}<=0.5,-1,IF(B{so_start+1}<=1,-2,-3)))))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('SO: Trend Adjustment')
ws[f'E{row}'], ws[f'F{row}'] = 'within 0.5ppt', '0'
row += 1

ws[f'A{row}'] = 'Pipeline Ratio'
ws[f'B{row}'] = f'=IF(B{so_start+4}<=0,"N/A",B{so_start+3}/B{so_start+4})'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},2),B{row})'
ws[f'D{row}'] = xlookup('SO: Pipeline Ratio')
ws[f'E{row}'], ws[f'F{row}'] = '0.5-2ppt above', '-1'
row += 1

ws[f'A{row}'] = 'Pipeline Adjustment'
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(C{row-1})),0,IF(C{row-1}<0.5,1,IF(C{row-1}<=1,0,IF(C{row-1}<=1.5,-1,-2))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('SO: Pipeline Adjustment')
ws[f'E{row}'], ws[f'F{row}'] = '>2ppt above', '-2'
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=MAX(1,MIN(10,5+C{row-4}+C{row-3}+C{row-1}))'
ws[f'C{row}'] = f'=B{row}'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = xlookup('SO: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = 'Base Score', '5'
final_score_rows['SO'] = row
row += 2

#===============================================================================
# MIGRATION / GDP GROWTH (3% weight)
#===============================================================================
row = add_header(row, 'MIGRATION / GDP GROWTH (3% weight)')
row = add_column_headers(row, ['Field', 'Formula / Value', 'Excel Calc', 'Python Calc', 'Scale', 'Score'])

mg_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Employment Growth - Market', xlookup('Employment Growth - Market')
ws[f'D{row}'] = xlookup('MG: Employment Growth - Market')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Employment Growth - US', xlookup('Employment Growth - US')
ws[f'D{row}'] = xlookup('MG: Employment Growth - US')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Pop Growth (5mi) %', xlookup('Population Growth % (5mi)')
ws[f'D{row}'] = xlookup('MG: Pop Growth (5mi)')
row += 1

ws[f'A{row}'] = 'Employment vs US'
ws[f'B{row}'] = f'=IF(OR(B{mg_start}="",B{mg_start+1}=""),"N/A",B{mg_start}-B{mg_start+1})'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},2),B{row})'
ws[f'D{row}'] = xlookup('MG: Employment vs US')
ws[f'E{row}'], ws[f'F{row}'] = '>=+1% vs US', 10
row += 1

ws[f'A{row}'] = 'Employment Score'
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(C{row-1})),"",IF(C{row-1}>=1,10,IF(C{row-1}>=0.5,8,IF(C{row-1}>=0,6,IF(C{row-1}>=-0.5,5,IF(C{row-1}>=-1,4,IF(C{row-1}>=-1.5,3,2)))))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('MG: Employment Score')
ws[f'E{row}'], ws[f'F{row}'] = '+0.5-1%', 8
row += 1

ws[f'A{row}'] = 'Population Score'
ws[f'B{row}'] = f'=IF(B{mg_start+2}="","",IF(B{mg_start+2}>=10,10,IF(B{mg_start+2}>=8,9,IF(B{mg_start+2}>=6,8,IF(B{mg_start+2}>=4,7,IF(B{mg_start+2}>=2,6,IF(B{mg_start+2}>=0,5,IF(B{mg_start+2}>=-2,4,2))))))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('MG: Population Score')
ws[f'E{row}'], ws[f'F{row}'] = '0-0.5%', 6
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(AND(C{row-2}="",C{row-1}=""),"",IF(C{row-2}="",C{row-1},IF(C{row-1}="",C{row-2},ROUND((C{row-2}+C{row-1})/2,0))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = xlookup('MG: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '<-1.5%', 2
final_score_rows['MG'] = row
row += 2

#===============================================================================
# PARKING RATIO (3% weight)
#===============================================================================
row = add_header(row, 'PARKING RATIO (3% weight)')
row = add_column_headers(row, ['Field', 'Formula / Value', 'Excel Calc', 'Python Calc', 'Scale', 'Score'])

pr_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Parking Ratio', xlookup('PR: Parking Ratio')
ws[f'D{row}'] = xlookup('PR: Parking Ratio')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Surface Spaces', xlookup('PR: Surface Spaces')
ws[f'D{row}'] = xlookup('PR: Surface Spaces')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Covered Spaces', xlookup('PR: Covered Spaces')
ws[f'D{row}'] = xlookup('PR: Covered Spaces')
row += 1

ws[f'A{row}'] = 'Base Score'
ws[f'B{row}'] = f'=IF(B{pr_start}="","",IF(B{pr_start}>=2,10,IF(B{pr_start}>=1.5,9,IF(B{pr_start}>=1.25,8,IF(B{pr_start}>=1,7,IF(B{pr_start}>=0.75,5,IF(B{pr_start}>=0.5,3,2)))))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('PR: Base Score')
ws[f'E{row}'], ws[f'F{row}'] = '>=2.0', 10
row += 1

ws[f'A{row}'] = 'Underground Penalty'
ws[f'B{row}'] = f'=IF(OR(B{pr_start+1}="",B{pr_start+2}=""),0,IF(AND(B{pr_start+1}=0,B{pr_start+2}>0),-1,0))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('PR: Underground Penalty')
ws[f'E{row}'], ws[f'F{row}'] = '1.5-2.0', 9
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(C{row-2}="","",MAX(1,MIN(10,C{row-2}+C{row-1})))'
ws[f'C{row}'] = f'=B{row}'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = xlookup('PR: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '1.25-1.5', 8
final_score_rows['PR'] = row
row += 1

# Additional scale rows
ws[f'E{row}'], ws[f'F{row}'] = '1.0-1.25', 7
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '0.75-1.0', 5
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '0.5-0.75', 3
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '<0.5', 2
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'Underground-only', -1
row += 2

#===============================================================================
# AMENITIES & LIFESTYLE (5% weight)
#===============================================================================
row = add_header(row, 'AMENITIES & LIFESTYLE (5% weight)')
row = add_column_headers(row, ['Field', 'Formula / Value', 'Excel Calc', 'Python Calc', 'Site Amenity', 'Points'])

am_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Site Score (max 5)', xlookup('AM: Site Score')
ws[f'D{row}'] = xlookup('AM: Site Score')
ws[f'E{row}'], ws[f'F{row}'] = 'Pool', 1.0
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Unit Score (max 5)', xlookup('AM: Unit Score')
ws[f'D{row}'] = xlookup('AM: Unit Score')
ws[f'E{row}'], ws[f'F{row}'] = 'Fitness Center', 1.0
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(AND(B{am_start}="",B{am_start+1}=""),"",MAX(1,MIN(10,B{am_start}+B{am_start+1})))'
ws[f'C{row}'] = f'=B{row}'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = xlookup('AM: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = 'Clubhouse', 1.0
final_score_rows['AM'] = row
row += 1

# More amenity points
ws[f'E{row}'], ws[f'F{row}'] = 'Dog Park', 0.5
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'Playground', 0.5
row += 1
ws[f'A{row}'] = 'Unit Amenity'
ws[f'B{row}'] = 'Points'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'].font = bold_font
ws[f'E{row}'], ws[f'F{row}'] = 'Business Center', 0.5
row += 1
ws[f'A{row}'], ws[f'B{row}'] = 'In-Unit W/D', 1.5
ws[f'E{row}'], ws[f'F{row}'] = 'Grill/Picnic', 0.25
row += 1
ws[f'A{row}'], ws[f'B{row}'] = 'W/D Hookup', 0.75
row += 1
ws[f'A{row}'], ws[f'B{row}'] = 'A/C', 1.0
row += 1
ws[f'A{row}'], ws[f'B{row}'] = 'Dishwasher', 0.5
row += 1
ws[f'A{row}'], ws[f'B{row}'] = 'Balcony/Patio', 0.5
row += 2

#===============================================================================
# UNIT MIX & SIZE (5% weight)
#===============================================================================
row = add_header(row, 'UNIT MIX & SIZE (5% weight)')
row = add_column_headers(row, ['Field', 'Formula / Value', 'Excel Calc', 'Python Calc', 'Scale', 'Score'])

um_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Total Units', xlookup('Number of Units')
ws[f'D{row}'] = xlookup('UM: Total Units')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Avg Unit Size (SF)', xlookup('Avg Unit Size (SF)')
ws[f'D{row}'] = xlookup('UM: Avg SF')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = '2BR Units', xlookup('Subject Units - 2BR')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = '3BR Units', xlookup('Subject Units - 3BR')
row += 1

ws[f'A{row}'] = '2-3BR Units'
ws[f'B{row}'] = f'=IF(OR(B{um_start+2}="",B{um_start+3}=""),"",B{um_start+2}+B{um_start+3})'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('UM: 2-3BR Units')
row += 1

ws[f'A{row}'] = '2-3BR %'
ws[f'B{row}'] = f'=IF(OR(C{row-1}="",B{um_start}=""),"",C{row-1}/B{um_start}*100)'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},1),B{row})'
ws[f'D{row}'] = xlookup('UM: 2-3BR %')
ws[f'E{row}'], ws[f'F{row}'] = '>=1000 SF', 10
row += 1

ws[f'A{row}'] = 'Size Score'
ws[f'B{row}'] = f'=IF(B{um_start+1}="","",IF(B{um_start+1}>=1000,10,IF(B{um_start+1}>=900,8,IF(B{um_start+1}>=800,6,IF(B{um_start+1}>=700,4,2)))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('UM: Size Score')
ws[f'E{row}'], ws[f'F{row}'] = '900-1000', 8
row += 1

ws[f'A{row}'] = 'Mix Score'
ws[f'B{row}'] = f'=IF(C{row-2}="","",IF(C{row-2}>=70,10,IF(C{row-2}>=60,8,IF(C{row-2}>=50,6,IF(C{row-2}>=40,4,3)))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'D{row}'] = xlookup('UM: Mix Score')
ws[f'E{row}'], ws[f'F{row}'] = '800-900', 6
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(AND(C{row-2}="",C{row-1}=""),"",IF(C{row-2}="",C{row-1},IF(C{row-1}="",C{row-2},ROUND((C{row-2}+C{row-1})/2,0))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = xlookup('UM: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '700-800', 4
final_score_rows['UM'] = row
row += 1

ws[f'E{row}'], ws[f'F{row}'] = '<700', 2
row += 1
ws[f'A{row}'] = 'Mix Scale'
ws[f'A{row}'].font = bold_font
row += 1
ws[f'A{row}'], ws[f'B{row}'] = '>=70% 2-3BR', 10
row += 1
ws[f'A{row}'], ws[f'B{row}'] = '60-70%', 8
row += 1
ws[f'A{row}'], ws[f'B{row}'] = '50-60%', 6
row += 1
ws[f'A{row}'], ws[f'B{row}'] = '40-50%', 4
row += 1
ws[f'A{row}'], ws[f'B{row}'] = '<40%', 3
row += 2

#===============================================================================
# LOSS-TO-LEASE (10% weight)
#===============================================================================
row = add_header(row, 'LOSS-TO-LEASE & NOI UPSIDE (10% weight)')
row = add_column_headers(row, ['Field', 'Formula / Value', 'Excel Calc', 'Python Calc', 'LTL %', 'Score'])

ltl_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Subject Rent (Avg)', xlookup('Subject Current Rent (Avg)')
ws[f'D{row}'] = xlookup('LTL: Subject Rent')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Comp Avg Rent', xlookup('Avg Comp Rent/Unit')
ws[f'D{row}'] = xlookup('LTL: Comp Avg Rent')
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Submarket Avg Rent', xlookup('Submarket Avg Rent')
ws[f'D{row}'] = xlookup('LTL: Submarket Rent')
row += 1

ws[f'A{row}'] = 'LTL vs Comps %'
ws[f'B{row}'] = f'=IF(OR(B{ltl_start}="",B{ltl_start+1}="",B{ltl_start+1}=0),"N/A",(B{ltl_start}-B{ltl_start+1})/B{ltl_start+1}*100)'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},1),B{row})'
ws[f'D{row}'] = xlookup('LTL: vs Comps %')
ws[f'E{row}'], ws[f'F{row}'] = '<=-20%', 10
row += 1

ws[f'A{row}'] = 'LTL vs Submarket %'
ws[f'B{row}'] = f'=IF(OR(B{ltl_start}="",B{ltl_start+2}="",B{ltl_start+2}=0),"N/A",(B{ltl_start}-B{ltl_start+2})/B{ltl_start+2}*100)'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},1),B{row})'
ws[f'D{row}'] = xlookup('LTL: vs Submarket %')
ws[f'E{row}'], ws[f'F{row}'] = '-15 to -20%', 9
row += 1

ws[f'A{row}'] = 'Blended LTL % (60/40)'
ws[f'B{row}'] = f'=IF(AND(ISNUMBER(C{row-2}),ISNUMBER(C{row-1})),C{row-2}*0.6+C{row-1}*0.4,IF(ISNUMBER(C{row-2}),C{row-2},IF(ISNUMBER(C{row-1}),C{row-1},"N/A")))'
ws[f'C{row}'] = f'=IF(ISNUMBER(B{row}),ROUND(B{row},1),B{row})'
ws[f'D{row}'] = xlookup('LTL: Blended %')
ws[f'E{row}'], ws[f'F{row}'] = '-10 to -15%', 8
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(C{row-1})),"",IF(C{row-1}<=-20,10,IF(C{row-1}<=-15,9,IF(C{row-1}<=-10,8,IF(C{row-1}<=-5,7,IF(C{row-1}<=-2.5,6,IF(C{row-1}<=2.5,5,IF(C{row-1}<=5,4,IF(C{row-1}<=10,3,IF(C{row-1}<=15,2,IF(C{row-1}<=20,1,0)))))))))))'
ws[f'C{row}'] = f'=B{row}'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = xlookup('LTL: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '-5 to -10%', 7
final_score_rows['LTL'] = row
row += 1

# Additional scale rows
ws[f'E{row}'], ws[f'F{row}'] = '-2.5 to -5%', 6
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '+/- 2.5%', 5
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '+2.5 to +5%', 4
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '+5 to +10%', 3
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '+10 to +15%', 2
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '+15 to +20%', 1
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '>+20%', 0
row += 2

#===============================================================================
# BUSINESS-FRIENDLY ENVIRONMENT (3% weight)
#===============================================================================
row = add_header(row, 'BUSINESS-FRIENDLY ENVIRONMENT (3% weight)')
row = add_column_headers(row, ['Field', 'Formula / Value', 'Excel Calc', 'Python Calc', 'States', 'Score'])

bf_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'State', xlookup('State')
ws[f'D{row}'] = xlookup('BF: State')
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
# State lookup formula
state_lookup = f'=LET(st,UPPER(B{bf_start}),IF(OR(st="TX",st="FL",st="TN",st="AZ"),10,IF(OR(st="GA",st="NC",st="SC",st="NV",st="IN"),9,IF(OR(st="KS",st="MO",st="OH",st="UT",st="OK",st="AL"),8,IF(OR(st="CO",st="ID",st="KY",st="AR",st="NE",st="LA",st="MS"),7,IF(OR(st="PA",st="MI",st="WI",st="VA",st="IA",st="MT",st="WY",st="SD",st="ND"),6,IF(OR(st="IL",st="MN",st="NM",st="WV",st="AK"),5,IF(OR(st="WA",st="MD",st="NH",st="DE"),4,IF(OR(st="MA",st="NJ",st="CT",st="HI",st="ME"),3,IF(OR(st="NY",st="VT",st="RI"),2,IF(OR(st="CA",st="OR",st="DC"),1,"")))))))))))'
ws[f'B{row}'] = state_lookup
ws[f'C{row}'] = f'=B{row}'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = xlookup('BF: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = 'TX,FL,TN,AZ', 10
final_score_rows['BF'] = row
row += 1

# State scale
ws[f'E{row}'], ws[f'F{row}'] = 'GA,NC,SC,NV,IN', 9
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'KS,MO,OH,UT,OK,AL', 8
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'CO,ID,KY,AR,NE,LA,MS', 7
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'PA,MI,WI,VA,IA...', 6
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'IL,MN,NM,WV,AK', 5
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'WA,MD,NH,DE', 4
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'MA,NJ,CT,HI,ME', 3
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'NY,VT,RI', 2
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'CA,OR,DC', 1
row += 2

# Summary section
row = add_header(row, 'FINAL SCORE SUMMARY')
row = add_column_headers(row, ['Score', 'Excel Formula Row', 'Excel Calc', 'Python Calc', 'Weight', 'Weighted'])

for name, r in final_score_rows.items():
    full_names = {
        'SD': 'Supply-Demand Drivers',
        'SO': 'Submarket Outlook',
        'MG': 'Migration/GDP',
        'PR': 'Parking Ratio',
        'AM': 'Amenities',
        'UM': 'Unit Mix & Size',
        'LTL': 'Loss-to-Lease',
        'BF': 'Business-Friendly'
    }
    weights = {
        'SD': '5%', 'SO': '10%', 'MG': '3%', 'PR': '3%',
        'AM': '5%', 'UM': '5%', 'LTL': '10%', 'BF': '3%'
    }
    weight_vals = {
        'SD': 0.05, 'SO': 0.10, 'MG': 0.03, 'PR': 0.03,
        'AM': 0.05, 'UM': 0.05, 'LTL': 0.10, 'BF': 0.03
    }
    ws[f'A{row}'] = full_names[name]
    ws[f'B{row}'] = f'Row {r}'
    ws[f'C{row}'] = f'=C{r}'
    ws[f'D{row}'] = f'=D{r}'
    ws[f'E{row}'] = weights[name]
    ws[f'F{row}'] = f'=IF(ISNUMBER(C{row}),C{row}*{weight_vals[name]},"")'
    row += 1

ws[f'A{row}'] = 'TOTAL WEIGHTED (44%)'
ws[f'A{row}'].font = bold_font
ws[f'C{row}'] = f'=SUM(F{row-8}:F{row-1})'
ws[f'C{row}'].fill = final_fill
ws[f'D{row}'] = '=SUM(D' + str(row-8) + ':D' + str(row-1) + ')*0.44/10'

# Save
wb.save(r'C:\Users\carso\OneDrive - UCB-O365\Work\RMP\Screener\RMP Screener_PreLinked_v3.xlsx')
print(f'Transparent Scoring Reference sheet complete!')
print(f'Final score rows: {final_score_rows}')
print(f'Total rows used: {row}')
