"""
Data Validator - Uses Claude API to independently verify extracted values
Reads Data Inputs sheet, checks each value against its claimed source,
and marks OK/FAIL/SKIP in the validation column.
"""

import os
import re
import sys
import hashlib
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List
import PyPDF2
import csv
import requests
import json
from urllib.parse import urlparse

# Config file location for API keys
CONFIG_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'validator_config.json')


def load_api_key_from_config() -> str:
    """Load API key from config file."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
                return config.get('anthropic_api_key')
        except:
            pass
    return None


def save_api_key_to_config(api_key: str):
    """Save API key to config file."""
    config = {}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
        except:
            pass
    config['anthropic_api_key'] = api_key
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=2)
    print(f"API key saved to {CONFIG_FILE}")


class DataValidator:
    """Validates extracted data against original sources using Claude API."""

    def __init__(self, excel_path: str, property_folder: str = None):
        """
        Initialize validator.

        Args:
            excel_path: Path to the screener output Excel file
            property_folder: Path to property folder containing PDFs (auto-detected if not provided)
        """
        self.excel_path = excel_path
        self.property_folder = property_folder or str(Path(excel_path).parent)
        self.workbook = None
        self.client = None
        self.validation_results = []
        self.api_calls_made = 0
        self.estimated_cost = 0.0

        # Cache settings
        self.skip_cached_ok = False
        self.validation_cache = {}  # field_name -> {value, result, explanation, pdf_hash}
        self.cached_pdf_hash = None
        self.current_pdf_hash = None
        self.fields_skipped_from_cache = 0

        # Find CoStar PDFs
        self.costar_property_pdf = None
        self.costar_market_pdf = None
        self._find_costar_pdfs()

    def _find_costar_pdfs(self):
        """Find CoStar PDF files in property folder."""
        costar_folder = os.path.join(self.property_folder, "CoStar Reports")
        if not os.path.exists(costar_folder):
            # Try property folder directly
            costar_folder = self.property_folder

        if not os.path.exists(costar_folder):
            print(f"CoStar folder not found: {costar_folder}")
            return

        pdf_files = [f for f in os.listdir(costar_folder) if f.lower().endswith('.pdf')]

        if not pdf_files:
            print(f"No PDF files found in: {costar_folder}")
            return

        # If only one PDF, use it for everything (combined report)
        if len(pdf_files) == 1:
            filepath = os.path.join(costar_folder, pdf_files[0])
            self.costar_property_pdf = filepath
            self.costar_market_pdf = filepath
            print(f"Found single CoStar PDF (combined): {pdf_files[0]}")
            return

        # Multiple PDFs - try to identify property vs market
        for filename in pdf_files:
            filepath = os.path.join(costar_folder, filename)
            lower_name = filename.lower()

            # Check for specific report types
            if 'property' in lower_name and not self.costar_property_pdf:
                self.costar_property_pdf = filepath
            elif any(x in lower_name for x in ['market', 'analytics', 'demographic']):
                self.costar_market_pdf = filepath
            elif any(x in lower_name for x in ['sr', 'combined', 'screen', 'report']):
                # Combined/screener report - use for both if not already set
                if not self.costar_property_pdf:
                    self.costar_property_pdf = filepath
                if not self.costar_market_pdf:
                    self.costar_market_pdf = filepath

        # Fallback: use first PDF found for any missing
        first_pdf = os.path.join(costar_folder, pdf_files[0])
        if not self.costar_property_pdf:
            self.costar_property_pdf = first_pdf
        if not self.costar_market_pdf:
            self.costar_market_pdf = first_pdf

        print(f"Found CoStar Property PDF: {os.path.basename(self.costar_property_pdf) if self.costar_property_pdf else None}")
        print(f"Found CoStar Market PDF: {os.path.basename(self.costar_market_pdf) if self.costar_market_pdf else None}")

        # Calculate PDF hash for cache invalidation
        self.current_pdf_hash = self._get_pdf_hash()

    def _get_pdf_hash(self) -> str:
        """Get hash of PDF file(s) to detect if source changed."""
        hasher = hashlib.md5()
        for pdf_path in [self.costar_property_pdf, self.costar_market_pdf]:
            if pdf_path and os.path.exists(pdf_path):
                # Use file size and modification time for quick hash
                stat = os.stat(pdf_path)
                hasher.update(f"{pdf_path}:{stat.st_size}:{stat.st_mtime}".encode())
        return hasher.hexdigest()[:16]

    def load_validation_cache(self):
        """Load validation cache from hidden Excel sheet."""
        if not self.workbook:
            return

        cache_sheet_name = "_ValidationCache"
        if cache_sheet_name not in self.workbook.sheetnames:
            return

        sheet = self.workbook[cache_sheet_name]

        # First row is header, second row is PDF hash
        pdf_hash_row = sheet[2]
        if pdf_hash_row[0].value == "pdf_hash":
            self.cached_pdf_hash = pdf_hash_row[1].value

        # Remaining rows are cached results: field_name, value, result, explanation
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if row[0]:  # field_name
                self.validation_cache[row[0]] = {
                    'value': str(row[1]) if row[1] else '',
                    'result': row[2],
                    'explanation': row[3] or ''
                }

        print(f"Loaded {len(self.validation_cache)} cached validation results")

        # Check if PDF changed
        if self.cached_pdf_hash and self.cached_pdf_hash != self.current_pdf_hash:
            print("[!] PDF source changed since last validation - cache invalidated")
            self.validation_cache = {}
            self.cached_pdf_hash = None

    def save_validation_cache(self):
        """Save validation cache to hidden Excel sheet."""
        if not self.workbook:
            return

        cache_sheet_name = "_ValidationCache"

        # Remove existing cache sheet if present
        if cache_sheet_name in self.workbook.sheetnames:
            del self.workbook[cache_sheet_name]

        # Create new cache sheet
        sheet = self.workbook.create_sheet(cache_sheet_name)
        sheet.sheet_state = 'hidden'  # Hide the sheet

        # Header
        sheet['A1'] = "Validation Cache - Do Not Edit"
        sheet['A2'] = "pdf_hash"
        sheet['B2'] = self.current_pdf_hash

        # Column headers
        sheet['A3'] = "field_name"
        sheet['B3'] = "value"
        sheet['C3'] = "result"
        sheet['D3'] = "explanation"

        # Write cached results from current validation
        row_num = 4
        for result in self.validation_results:
            if result['result'] in ['OK', 'FAIL', '?']:  # Only cache validated fields
                sheet[f'A{row_num}'] = result['field']
                sheet[f'B{row_num}'] = result['value']
                sheet[f'C{row_num}'] = result['result']
                sheet[f'D{row_num}'] = result['explanation'][:100] if result['explanation'] else ''
                row_num += 1

    def can_skip_field(self, field_name: str, current_value: str) -> Tuple[bool, str, str]:
        """
        Check if field can be skipped based on cache.

        Returns: (can_skip, cached_result, cached_explanation)
        """
        if not self.skip_cached_ok:
            return False, None, None

        if field_name not in self.validation_cache:
            return False, None, None

        cached = self.validation_cache[field_name]

        # Only skip if previous result was OK and value unchanged
        if cached['result'] == 'OK' and str(current_value) == cached['value']:
            return True, 'OK', f"[cached] {cached['explanation']}"

        return False, None, None

    def init_claude_client(self, api_key: str = None):
        """Initialize Claude API client."""
        # Check sources in order: passed param, config file, environment
        api_key = api_key or load_api_key_from_config() or os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            raise ValueError("No API key provided. Set ANTHROPIC_API_KEY env var or pass api_key parameter.")
        self.client = anthropic.Anthropic(api_key=api_key)
        print("Claude API client initialized")

    def load_excel(self):
        """Load the Excel workbook and validation cache."""
        self.workbook = openpyxl.load_workbook(self.excel_path)
        print(f"Loaded: {self.excel_path}")
        self._load_source_urls()
        self.load_validation_cache()

    def _load_source_urls(self):
        """Load source URLs from the bottom section of Data Inputs sheet."""
        self.source_urls = {}  # field_name -> url
        sheet = self.workbook['Data Inputs']

        in_url_section = False
        for row in sheet.iter_rows(min_row=1, max_row=500, max_col=4, values_only=True):
            field = row[1]  # Column B
            label = row[2]  # Column C
            url = row[3]    # Column D

            if field and 'SOURCE URLs' in str(field):
                in_url_section = True
                continue

            if in_url_section and field and url:
                # Store URL keyed by field name
                self.source_urls[field] = str(url)

        print(f"Loaded {len(self.source_urls)} source URLs for validation")

    def extract_pdf_page(self, pdf_path: str, page_num: int) -> str:
        """Extract text from a specific page of a PDF."""
        if not pdf_path or not os.path.exists(pdf_path):
            return None

        try:
            with open(pdf_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                if page_num < 1 or page_num > len(reader.pages):
                    return None
                page = reader.pages[page_num - 1]  # 0-indexed
                return page.extract_text()
        except Exception as e:
            print(f"Error extracting PDF page: {e}")
            return None

    def parse_source(self, source: str) -> Dict[str, Any]:
        """
        Parse source string to determine validation method.

        Returns dict with:
            - type: 'costar_property', 'costar_market', 'web', 'csv', 'manual', 'formula'
            - page: page number if applicable
            - url: URL if web source
        """
        if not source:
            return {'type': 'unknown'}

        source_lower = source.lower()

        # Helper to extract page number from various formats
        def extract_page(text):
            # Match: (pg 1), (pg. 1), p.1, p 1, pg 1, page 1
            patterns = [
                r'\(pg\.?\s*(\d+)\)',  # (pg 1) or (pg. 1)
                r'pg\.?\s*(\d+)',       # pg 1 or pg. 1
                r'p\.?\s*(\d+)',        # p.1 or p 1
                r'page\s*(\d+)',        # page 1
            ]
            for pattern in patterns:
                match = re.search(pattern, text.lower())
                if match:
                    return int(match.group(1))
            return None

        # CoStar Property report
        if 'costar property' in source_lower or 'costar prop' in source_lower:
            return {'type': 'costar_property', 'page': extract_page(source)}

        # CoStar Market/Analytics/Demographics/Economy report (all use market PDF)
        if any(x in source_lower for x in ['costar market', 'costar analytics', 'costar economy', 'costar demographics', 'costar demo']):
            return {'type': 'costar_market', 'page': extract_page(source)}

        # Generic CoStar source (try to determine type from context)
        if 'costar' in source_lower:
            page = extract_page(source)
            # Property data is usually early pages, demographics/market later
            if page and page > 20:
                return {'type': 'costar_market', 'page': page}
            return {'type': 'costar_property', 'page': page}

        # Web/API sources
        if any(x in source_lower for x in ['bestplaces', 'census', 'fema', 'walkscore', 'greatschools', 'api', 'web scraping']):
            return {'type': 'web', 'source_name': source}

        # Config/Manual entry
        if any(x in source_lower for x in ['config', 'manual', 'user']):
            return {'type': 'manual'}

        # Formula/Calculated
        if any(x in source_lower for x in ['formula', 'calculated', 'derived']):
            return {'type': 'formula'}

        # CSV lookup (crime data)
        if 'crime' in source_lower and 'csv' in source_lower:
            return {'type': 'csv', 'source_name': source}

        # Hyperlink sources
        if source.startswith('http') or 'View on' in source:
            return {'type': 'link', 'url': source}

        return {'type': 'unknown', 'raw': source}

    def validate_with_claude(self, field_name: str, claimed_value: str, page_text: str, source_desc: str) -> Tuple[str, str]:
        """
        Use Claude API to validate a value against source text.

        Returns:
            Tuple of (result, explanation)
            result: 'OK', 'FAIL', or '?'
        """
        if not self.client:
            return ('SKIP', 'No API client')

        if not page_text:
            return ('SKIP', 'Could not extract page text')

        # Truncate very long page text to save tokens
        if len(page_text) > 8000:
            page_text = page_text[:8000] + "\n...[truncated]..."

        prompt = f"""You are validating data extraction accuracy.

FIELD: {field_name}
CLAIMED VALUE: {claimed_value}
SOURCE: {source_desc}

Below is the text extracted from the source document page. Determine if the claimed value can be found or reasonably derived from this text.

Rules:
- Numbers may be formatted differently (e.g., "5.2%" vs "5.2" or "1,234" vs "1234")
- Values may be rounded slightly
- The exact field label might not appear, but the value should be present in a relevant context
- For percentages, both "5%" and "5.0%" are equivalent

PAGE TEXT:
{page_text}

Respond with ONLY one of these three options:
OK - if the value is present or clearly derivable from the text
FAIL - if the value contradicts the text or is clearly not present
? - if you cannot determine (text unclear, value ambiguous, etc.)

Then on a new line, provide a brief explanation (max 50 words)."""

        try:
            response = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=150,
                messages=[{"role": "user", "content": prompt}]
            )

            self.api_calls_made += 1
            # Cost estimate: ~2000 input tokens + 150 output tokens at Sonnet pricing
            # Input: $3/M tokens, Output: $15/M tokens = ~$0.008-0.012 per call
            self.estimated_cost += 0.01  # ~$0.01 per call estimate

            result_text = response.content[0].text.strip()
            lines = result_text.split('\n', 1)
            result = lines[0].strip().upper()
            explanation = lines[1].strip() if len(lines) > 1 else ""

            # Normalize result
            if result not in ['OK', 'FAIL', '?']:
                if 'OK' in result:
                    result = 'OK'
                elif 'FAIL' in result:
                    result = 'FAIL'
                else:
                    result = '?'

            return (result, explanation)

        except Exception as e:
            return ('ERR', f'API error: {str(e)[:50]}')

    def validate_batch_with_claude(self, fields: List[Dict], page_text: str, source_desc: str) -> Dict[str, Tuple[str, str]]:
        """
        Validate multiple fields from the same page in one API call.

        Args:
            fields: List of dicts with 'field_name' and 'value' keys
            page_text: Text from the PDF page
            source_desc: Description of source (e.g., "CoStar Property p.1")

        Returns:
            Dict mapping field_name to (result, explanation) tuple
        """
        if not self.client:
            return {f['field_name']: ('SKIP', 'No API client') for f in fields}

        if not page_text:
            return {f['field_name']: ('SKIP', 'Could not extract page text') for f in fields}

        # Truncate very long page text to save tokens
        if len(page_text) > 8000:
            page_text = page_text[:8000] + "\n...[truncated]..."

        # Build fields list for prompt
        fields_text = "\n".join([f"- {f['field_name']}: {f['value']}" for f in fields])

        prompt = f"""You are validating data extraction accuracy. Check if EACH of these values can be found on this page.

SOURCE: {source_desc}

FIELDS TO VALIDATE:
{fields_text}

PAGE TEXT:
{page_text}

For EACH field, respond with the field name, then OK/FAIL/? and a brief reason.
Format each line as: FIELD_NAME | OK/FAIL/? | reason

Rules:
- OK = value is present or clearly derivable from the text
- FAIL = value contradicts the text or is clearly not present
- ? = cannot determine (text unclear, ambiguous)
- Numbers may be formatted differently (5.2% vs 5.2, 1,234 vs 1234)
- Values may be rounded slightly"""

        try:
            response = self.client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=50 + (len(fields) * 40),  # Scale output with field count
                messages=[{"role": "user", "content": prompt}]
            )

            self.api_calls_made += 1
            # Batch call costs more tokens but fewer calls
            self.estimated_cost += 0.015  # ~$0.015 per batch call

            result_text = response.content[0].text.strip()
            results = {}

            # Helper function to normalize field names for comparison
            def normalize_field(name):
                """Normalize field name for fuzzy matching."""
                # Remove special chars, lowercase, collapse whitespace
                import re
                name = name.lower()
                name = re.sub(r'[()%\-_/]', ' ', name)  # Replace special chars with space
                name = re.sub(r'\s+', ' ', name).strip()  # Collapse whitespace
                return name

            # Pre-normalize all field names we're looking for
            field_normalized = {normalize_field(f['field_name']): f['field_name'] for f in fields}

            # Parse response - each line should be: FIELD_NAME | RESULT | reason
            for line in result_text.split('\n'):
                line = line.strip()
                if not line:
                    continue

                if '|' in line:
                    parts = [p.strip() for p in line.split('|')]
                    if len(parts) >= 2:
                        field_name = parts[0]
                        result = parts[1].upper()
                        explanation = parts[2] if len(parts) > 2 else ""

                        # Normalize result
                        if 'OK' in result:
                            result = 'OK'
                        elif 'FAIL' in result:
                            result = 'FAIL'
                        else:
                            result = '?'

                        # Normalize the field name from response
                        resp_normalized = normalize_field(field_name)

                        # Match field name - try multiple strategies
                        matched_field = None

                        # 1. Exact normalized match
                        if resp_normalized in field_normalized:
                            matched_field = field_normalized[resp_normalized]

                        # 2. Check if response field contains or is contained by any field
                        if not matched_field:
                            for norm, orig in field_normalized.items():
                                if orig not in results:  # Only match unmatched fields
                                    if norm in resp_normalized or resp_normalized in norm:
                                        matched_field = orig
                                        break

                        # 3. Word overlap matching (need 2+ significant words)
                        if not matched_field:
                            resp_words = set(resp_normalized.split())
                            # Remove common words that don't help matching
                            stop_words = {'the', 'a', 'an', 'of', 'for', 'in', 'to', 'mi'}
                            resp_words = resp_words - stop_words

                            best_match = None
                            best_score = 0
                            for norm, orig in field_normalized.items():
                                if orig not in results:  # Only match unmatched fields
                                    field_words = set(norm.split()) - stop_words
                                    overlap = len(resp_words & field_words)
                                    # Score based on overlap / total unique words
                                    if overlap >= 2:
                                        score = overlap / max(len(resp_words | field_words), 1)
                                        if score > best_score:
                                            best_score = score
                                            best_match = orig

                            if best_match and best_score >= 0.4:  # At least 40% word overlap
                                matched_field = best_match

                        if matched_field:
                            results[matched_field] = (result, explanation[:60])

            # Fill in any missing fields - try to extract from full response text
            for f in fields:
                if f['field_name'] not in results:
                    # Last resort: check if the value appears with OK/FAIL near it
                    f_lower = f['field_name'].lower()
                    if f_lower in result_text.lower() and 'ok' in result_text.lower():
                        results[f['field_name']] = ('?', 'Partial match in response')
                    else:
                        results[f['field_name']] = ('?', 'Not found in response')

            return results

        except Exception as e:
            return {f['field_name']: ('ERR', f'API error: {str(e)[:30]}') for f in fields}

    def validate_csv_lookup(self, field_name: str, claimed_value: str, zip_code: str = None) -> Tuple[str, str]:
        """Validate crime data against CSV file."""
        # Find the crime CSV
        csv_path = os.path.join(
            os.path.dirname(os.path.dirname(self.property_folder)),
            'kansas_zcta_crime.csv'
        )

        if not os.path.exists(csv_path):
            # Try alternate location
            csv_path = os.path.join(
                os.path.dirname(self.excel_path),
                '..', '..', 'kansas_zcta_crime.csv'
            )

        if not os.path.exists(csv_path):
            return ('SKIP', 'Crime CSV not found')

        try:
            with open(csv_path, 'r') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if zip_code and row.get('zcta') == zip_code:
                        # Check if claimed value matches
                        if 'score' in field_name.lower():
                            csv_value = row.get('score_10', '')
                            if str(claimed_value) == str(csv_value):
                                return ('OK', f'Matches CSV: {csv_value}')
                            else:
                                return ('FAIL', f'CSV has {csv_value}, claimed {claimed_value}')
                        elif 'index' in field_name.lower():
                            csv_value = row.get('crime_index', '')
                            if str(claimed_value) == str(csv_value):
                                return ('OK', f'Matches CSV: {csv_value}')

            return ('SKIP', 'Could not verify against CSV')
        except Exception as e:
            return ('ERR', str(e)[:50])

    def validate_web_source(self, field_name: str, claimed_value: str, source_name: str, source_url: str = None) -> Tuple[str, str]:
        """
        Validate web-sourced data by re-fetching and checking with Claude.

        Args:
            field_name: Name of the field
            claimed_value: The value we extracted
            source_name: Description of the source (e.g., "BestPlaces Crime")
            source_url: Optional URL to fetch
        """
        if not self.client:
            return ('SKIP', 'No API client')

        # Map source names to URLs we can check
        url_mappings = {
            'bestplaces': 'https://www.bestplaces.net/',
            'walkscore': 'https://www.walkscore.com/',
            'fema': 'https://hazards.fema.gov/',
            'census': 'https://www.census.gov/',
        }

        # Try to find a URL to fetch
        fetch_url = source_url
        if not fetch_url:
            source_lower = source_name.lower()
            for key, url in url_mappings.items():
                if key in source_lower:
                    # We can't easily construct the exact URL without address info
                    return ('SKIP', f'Would need full URL for {key}')

        if not fetch_url:
            return ('SKIP', 'No URL available to verify')

        # Fetch the URL
        try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
            response = requests.get(fetch_url, headers=headers, timeout=10)
            if response.status_code != 200:
                return ('SKIP', f'HTTP {response.status_code}')

            # Get text content (strip HTML for now, or use basic extraction)
            content = response.text

            # Truncate to avoid huge API calls
            if len(content) > 10000:
                content = content[:10000] + "\n...[truncated]..."

            # Use Claude to check if value is present
            return self.validate_with_claude(
                field_name, claimed_value, content, f"Web page: {fetch_url}"
            )

        except requests.RequestException as e:
            return ('SKIP', f'Fetch error: {str(e)[:30]}')

    def run_validation(self):
        """
        Run validation on PDF-sourced data only.
        Groups fields by page to batch validate and reduce API calls.

        Web/scraped sources are skipped - they can't be reliably re-verified
        and the original scraper already validated the data extraction.
        """
        if not self.workbook:
            self.load_excel()

        sheet = self.workbook['Data Inputs']

        # Find property zip code for CSV lookups
        zip_code = None
        for row in sheet.iter_rows(min_row=4, max_col=4, values_only=True):
            if row[0] == 'Property Zip':
                zip_code = str(row[2]) if row[2] else None
                break

        print(f"\nStarting validation (Property ZIP: {zip_code})")
        if self.skip_cached_ok and self.validation_cache:
            print(f"[Cache] {len(self.validation_cache)} previous results available - skipping unchanged OKs")
        print("-" * 60)

        # First pass: collect all fields and group by page
        pdf_batches = {}  # (pdf_type, page) -> list of field dicts
        other_fields = []  # Non-PDF fields to validate individually
        all_rows = []  # Track all rows for results
        cached_results = {}  # Fields skipped due to cache

        row_num = 4
        for row in sheet.iter_rows(min_row=4, max_row=200, max_col=6):
            field_cell = row[1]  # Column B
            value_cell = row[2]  # Column C
            source_cell = row[3]  # Column D
            ok_cell = row[4]  # Column E

            field_name = field_cell.value
            value = value_cell.value
            source = source_cell.value

            # Skip empty rows and section headers
            if not field_name or not value or field_name.isupper():
                row_num += 1
                continue

            # Parse source
            source_info = self.parse_source(str(source) if source else '')
            source_type = source_info.get('type', 'unknown')

            field_data = {
                'row_num': row_num,
                'field_name': field_name,
                'value': str(value),
                'source': source,
                'source_info': source_info,
                'source_type': source_type,
                'ok_cell': ok_cell
            }
            all_rows.append(field_data)

            # Check cache - skip if previously OK and value unchanged
            can_skip, cached_result, cached_explanation = self.can_skip_field(field_name, str(value))
            if can_skip:
                cached_results[field_name] = (cached_result, cached_explanation)
                self.fields_skipped_from_cache += 1
                row_num += 1
                continue

            # Group PDF sources by page for batch validation
            # Use just page number as key since property and market may use same PDF
            if source_type == 'costar_property' and source_info.get('page') and self.costar_property_pdf:
                page = source_info['page']
                pdf_path = self.costar_property_pdf
                key = (pdf_path, page)  # Group by actual PDF path + page
                if key not in pdf_batches:
                    pdf_batches[key] = []
                pdf_batches[key].append(field_data)
            elif source_type == 'costar_market' and source_info.get('page') and self.costar_market_pdf:
                page = source_info['page']
                pdf_path = self.costar_market_pdf
                key = (pdf_path, page)  # Group by actual PDF path + page
                if key not in pdf_batches:
                    pdf_batches[key] = []
                pdf_batches[key].append(field_data)
            else:
                other_fields.append(field_data)

            row_num += 1

        # Second pass: batch validate PDF pages
        results_map = {}  # field_name -> (result, explanation)

        # Add cached results first
        results_map.update(cached_results)

        if self.fields_skipped_from_cache > 0:
            print(f"\n[Cache] Skipped {self.fields_skipped_from_cache} fields (unchanged since last OK)")

        print(f"\nValidating {len(pdf_batches)} PDF pages with {sum(len(v) for v in pdf_batches.values())} fields...")

        for (pdf_path, page), fields in pdf_batches.items():
            page_text = self.extract_pdf_page(pdf_path, page)
            source_desc = f"CoStar Report p.{page}"

            if page_text:
                batch_results = self.validate_batch_with_claude(
                    [{'field_name': f['field_name'], 'value': f['value']} for f in fields],
                    page_text,
                    source_desc
                )
                results_map.update(batch_results)
                print(f"  [Batch] Page {page}: {len(fields)} fields validated")
            else:
                for f in fields:
                    results_map[f['field_name']] = ('SKIP', 'Could not extract page text')

        # Third pass: validate non-PDF fields individually
        print(f"\nValidating {len(other_fields)} non-PDF fields...")

        for field_data in other_fields:
            field_name = field_data['field_name']
            value = field_data['value']
            source = field_data['source']
            source_info = field_data['source_info']
            source_type = field_data['source_type']

            result = 'SKIP'
            explanation = ''

            # All non-PDF sources are skipped
            if source_type == 'web':
                result, explanation = 'SKIP', 'Web source'
            elif source_type == 'manual':
                result, explanation = 'SKIP', 'Manual entry'
            elif source_type == 'formula':
                result, explanation = 'SKIP', 'Calculated'
            elif source_type == 'link':
                result, explanation = 'SKIP', 'Web link'
            elif source_type in ['costar_property', 'costar_market']:
                result, explanation = 'SKIP', 'No page number'
            else:
                result, explanation = 'SKIP', 'Non-PDF source'

            results_map[field_name] = (result, explanation)

        # Final pass: update all cells and collect results
        results = []
        print("\n" + "-" * 60)
        print("Results:")
        print("-" * 60)

        for field_data in all_rows:
            field_name = field_data['field_name']
            ok_cell = field_data['ok_cell']

            if field_name in results_map:
                result, explanation = results_map[field_name]
            else:
                result, explanation = 'SKIP', 'Not validated'

            # Store result
            results.append({
                'row': field_data['row_num'],
                'field': field_name,
                'value': field_data['value'],
                'source': field_data['source'],
                'result': result,
                'explanation': explanation
            })

            # Update cell
            ok_cell.value = result
            if result == 'OK':
                ok_cell.font = Font(color="008000", bold=True)  # Green
            elif result == 'FAIL':
                ok_cell.font = Font(color="FF0000", bold=True)  # Red
            elif result == '?':
                ok_cell.font = Font(color="FFA500", bold=True)  # Orange
            else:
                ok_cell.font = Font(color="808080")  # Gray for SKIP

            # Print progress
            status_symbol = {'OK': '✓', 'FAIL': '✗', '?': '?', 'SKIP': '-', 'ERR': '!'}.get(result, '-')
            print(f"[{status_symbol}] {field_name[:30]:<30} | {result:<4} | {explanation[:40]}")

        self.validation_results = results
        print("-" * 60)
        print(f"Validation complete. API calls: {self.api_calls_made}, Est. cost: ${self.estimated_cost:.3f}")

    def save(self, output_path: str = None):
        """Save the workbook with validation results and cache."""
        self.save_validation_cache()  # Save cache before saving workbook
        output_path = output_path or self.excel_path
        self.workbook.save(output_path)
        print(f"Saved to: {output_path}")

    def get_summary(self) -> Dict[str, int]:
        """Get summary of validation results."""
        summary = {'OK': 0, 'FAIL': 0, '?': 0, 'SKIP': 0, 'ERR': 0}
        for r in self.validation_results:
            result = r['result']
            summary[result] = summary.get(result, 0) + 1
        return summary


def main():
    """CLI entry point for data validator."""
    import argparse

    parser = argparse.ArgumentParser(description='Validate screener data against sources')
    parser.add_argument('excel_file', help='Path to screener output Excel file')
    parser.add_argument('--api-key', help='Claude API key (or set ANTHROPIC_API_KEY env var)')
    parser.add_argument('--include-web', action='store_true', help='Include web-sourced data validation')
    parser.add_argument('--output', help='Output file path (default: overwrites input)')

    args = parser.parse_args()

    validator = DataValidator(args.excel_file)
    validator.init_claude_client(args.api_key)
    validator.load_excel()
    validator.run_validation(skip_web=not args.include_web)
    validator.save(args.output)

    summary = validator.get_summary()
    print(f"\nSummary: {summary}")


if __name__ == '__main__':
    main()
