"""
Excel Writer - Populates the screener Excel file via Data Inputs sheet
All data goes to a single Data Inputs tab, other sheets reference it via formulas
"""

import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Tuple, Any, Dict
import shutil
import re
from datetime import datetime

from data_inputs_mapper import DataInputsMapper, get_formula_mappings
from score_calculator import ScoreCalculator


class ScreenerExcelWriter:
    """Writes data to Excel screener file via Data Inputs sheet."""

    def __init__(self, source_file: str, output_file: str):
        """
        Initialize writer with file paths.

        Args:
            source_file: Path to source screener Excel file
            output_file: Path where updated file should be saved
        """
        self.source_file = source_file
        self.output_file = output_file
        self.workbook = None
        self.changes_made = []
        self.mapper = DataInputsMapper()

    def load_workbook(self):
        """Load the Excel workbook."""
        # Make a copy of source file
        shutil.copy2(self.source_file, self.output_file)

        # Load the copied workbook
        self.workbook = load_workbook(self.output_file)
        print(f"Loaded workbook: {self.output_file}")
        print(f"Sheets: {self.workbook.sheetnames}")

    def create_data_inputs_sheet(self, extracted_data: Dict[str, Any], config: Dict[str, Any]):
        """
        Create and populate the Data Inputs sheet.

        Args:
            extracted_data: All extracted data from various sources
            config: Property configuration
        """
        if not self.workbook:
            raise ValueError("Workbook not loaded. Call load_workbook() first.")

        # Create or get the Data Inputs sheet
        sheet_name = "Data Inputs"
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
        else:
            # Insert at the beginning
            sheet = self.workbook.create_sheet(sheet_name, 0)

        # Define styles
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set column widths
        sheet.column_dimensions['A'].width = 3
        sheet.column_dimensions['B'].width = 28
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 22
        sheet.column_dimensions['E'].width = 8
        sheet.column_dimensions['F'].width = 55  # Description column

        # Clear existing content
        for row in sheet.iter_rows(min_row=1, max_row=100, max_col=6):
            for cell in row:
                cell.value = None
                cell.fill = PatternFill()
                cell.font = Font()

        # Title
        sheet['B1'] = 'DATA INPUTS'
        sheet['B1'].font = header_font
        sheet['B2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

        # Column headers
        for col, header in [('B', 'Field'), ('C', 'Value'), ('D', 'Source'), ('E', 'OK?'), ('F', 'Description')]:
            cell = sheet[f'{col}3']
            cell.value = header
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Calculate Stage 2 scores
        score_calc = ScoreCalculator()
        extracted_data = score_calc.calculate_all_scores(extracted_data)

        # Get mapped data
        updates = self.mapper.map_to_data_inputs(extracted_data, config)

        # Get PDF path for hyperlinks
        source_pdf = extracted_data.get('_source_pdf', {})
        pdf_path = source_pdf.get('full_path')

        # Track which rows have data for populating (now includes description)
        row_data = {row: (field, value, source, desc) for row, field, value, source, desc in updates}

        # Populate sheet section by section
        current_row = 4
        current_section = None

        section_titles = {
            'property': 'PROPERTY INFO',
            'location': 'LOCATION',
            'demo_1mi': 'DEMOGRAPHICS - 1 MILE RADIUS',
            'demo_3mi': 'DEMOGRAPHICS - 3 MILE RADIUS',
            'demo_5mi': 'DEMOGRAPHICS - 5 MILE RADIUS',
            'demo_other': 'DEMOGRAPHICS - OTHER',
            'market': 'MARKET DATA',
            'submarket': 'SUBMARKET DATA',
            'employment': 'EMPLOYMENT DATA',
            'scores': 'SCORES & RATINGS',
            'rent_subject': 'RENT DATA - SUBJECT',
            'rent_market': 'RENT DATA - MARKET',
            'stage2_scores': 'STAGE 2 CALCULATED SCORES',
            'comps': 'COMPS SUMMARY',
        }

        for field_def in self.mapper.FIELD_DEFINITIONS:
            # Handle both 3-element and 4-element tuples
            field_name = field_def[0]
            section = field_def[2]

            # Add section header if new section
            if section != current_section:
                current_row += 1  # Blank row
                sheet[f'B{current_row}'] = section_titles.get(section, section.upper())
                sheet[f'B{current_row}'].font = section_font
                sheet[f'B{current_row}'].fill = section_fill
                # Merge section header across columns
                for col in ['C', 'D', 'E', 'F']:
                    sheet[f'{col}{current_row}'].fill = section_fill
                current_section = section
                current_row += 1

            # Add field row
            sheet[f'B{current_row}'] = field_name
            sheet[f'B{current_row}'].border = thin_border

            # Add value, source, and description if we have data
            if current_row in row_data:
                _, value, source, description = row_data[current_row]

                # Special handling for URLs - make them clickable
                if field_name in ['Crime Lookup URL'] and value and str(value).startswith('http'):
                    sheet[f'C{current_row}'] = 'View on BestPlaces'
                    sheet[f'C{current_row}'].hyperlink = value
                    sheet[f'C{current_row}'].font = Font(color="0563C1", underline="single")
                elif field_name == 'Manual School Lookup' and value and str(value).startswith('http'):
                    # Put link in Sources column (D) instead of Values (C)
                    sheet[f'C{current_row}'] = ''  # Leave value empty
                    sheet[f'D{current_row}'] = 'View on GreatSchools'
                    sheet[f'D{current_row}'].hyperlink = value
                    sheet[f'D{current_row}'].font = Font(color="0563C1", underline="single")
                    source = None  # Don't overwrite source below
                else:
                    sheet[f'C{current_row}'] = value

                # Source column - page numbers shown for manual PDF lookup
                # Handle source as either string or dict with URL
                # Skip if source was already handled above (e.g., Manual School Lookup)
                if source is not None:
                    if isinstance(source, dict) and source.get('url'):
                        # Create clickable hyperlink
                        sheet[f'D{current_row}'] = source.get('label', 'Source')
                        sheet[f'D{current_row}'].hyperlink = source['url']
                        sheet[f'D{current_row}'].font = Font(color="0563C1", underline="single")
                    else:
                        sheet[f'D{current_row}'] = source

                # Description column (F) - explains what the field means/how calculated
                if description:
                    sheet[f'F{current_row}'] = description
                    sheet[f'F{current_row}'].font = Font(size=9, italic=True, color="666666")

                self.changes_made.append({
                    'sheet': sheet_name,
                    'field': field_name,
                    'value': value,
                    'source': source.get('label', source) if isinstance(source, dict) else source
                })

            # Apply borders to all columns
            for col in ['C', 'D', 'E', 'F']:
                sheet[f'{col}{current_row}'].border = thin_border

            current_row += 1

        # === ADD SOURCE URLs SECTION AT BOTTOM ===
        # Collect all URLs from sources for validator reference
        source_urls = []
        for row, field, value, source, desc in updates:
            # Skip manual lookup fields
            if field == 'Manual School Lookup':
                continue
            if isinstance(source, dict) and source.get('url'):
                source_urls.append((field, source.get('label', ''), source['url']))
            elif isinstance(value, str) and value.startswith('http'):
                source_urls.append((field, 'Direct URL', value))

        # Also add scraped data URLs
        scraped = extracted_data.get('scraped', {})
        if scraped.get('flood_source_url'):
            source_urls.append(('Flood Zone', 'FEMA NFHL', scraped['flood_source_url']))
        if scraped.get('greatschools_url'):
            source_urls.append(('School Ratings', 'GreatSchools', scraped['greatschools_url']))

        if source_urls:
            current_row += 2  # Extra spacing
            sheet[f'B{current_row}'] = 'SOURCE URLs (for validation)'
            sheet[f'B{current_row}'].font = section_font
            sheet[f'B{current_row}'].fill = section_fill
            for col in ['C', 'D', 'E', 'F']:
                sheet[f'{col}{current_row}'].fill = section_fill
            current_row += 1

            for field_name, label, url in source_urls:
                sheet[f'B{current_row}'] = field_name
                sheet[f'C{current_row}'] = label
                sheet[f'D{current_row}'] = url  # Plain text URL for validator
                sheet[f'D{current_row}'].font = Font(size=9, color="666666")
                for col in ['B', 'C', 'D']:
                    sheet[f'{col}{current_row}'].border = thin_border
                current_row += 1

        print(f"[OK] Created Data Inputs sheet with {len(updates)} values populated")

    def add_reference_formulas(self):
        """Add formulas to other sheets that reference Data Inputs."""
        if not self.workbook:
            return

        formulas = get_formula_mappings()

        for (sheet_name, cell), formula in formulas.items():
            if sheet_name in self.workbook.sheetnames:
                try:
                    self.workbook[sheet_name][cell] = formula
                except Exception as e:
                    print(f"[WARNING] Could not add formula to {sheet_name}!{cell}: {e}")

        print(f"[OK] Added {len(formulas)} reference formulas to other sheets")

    def write_rent_comps(self, extracted_data: Dict[str, Any], config: Dict[str, Any] = None):
        """
        Write rent comparables to the Rent Comps sheet.
        Subject property goes first, then comps.

        Excel columns (Row 8 headers, data starts row 9):
        B: Group, C: Map #, D: Building Name, E: Address, F: City, G: State,
        H: Rating, I: Apartments.com Ad Level, J: Common View, K: % Overlap,
        L: Units, M: Stories, N: Yr Blt/Ren, O: Avg SF, P: mi Away,
        Q: Rent/SF, R: Rent/Unit, S: Studio, T: 1 Beds, U: 2 Beds, V: 3 Beds,
        W: Occ %, X: Concess %, Y: #Studio, Z: #1 Beds, AA: #2 Beds, AB: #3 Beds, AC: Neighborhood
        """
        if not self.workbook:
            return

        if 'Rent Comps' not in self.workbook.sheetnames:
            return

        sheet = self.workbook['Rent Comps']
        rent_comps = extracted_data.get('rent_comps', {})
        comps = rent_comps.get('comparable_properties', [])
        prop = extracted_data.get('property', {})

        # Data starts at row 9 (row 8 is headers)
        start_row = 9
        current_row = start_row

        # === SUBJECT PROPERTY FIRST (Row 9) ===
        sheet[f'B{current_row}'] = 'Subject'

        # Get property name from config if available
        if config and config.get('property_name'):
            sheet[f'D{current_row}'] = config['property_name']

        # Get address info from config
        if config and config.get('property_details'):
            details = config['property_details']
            if details.get('address'):
                sheet[f'E{current_row}'] = details['address']
            if details.get('city'):
                sheet[f'F{current_row}'] = details['city']
            if details.get('state'):
                sheet[f'G{current_row}'] = details['state']

        # Subject property data
        if prop.get('units'):
            sheet[f'L{current_row}'] = prop['units']
        if prop.get('stories'):
            sheet[f'M{current_row}'] = prop['stories']
        if prop.get('vintage'):
            sheet[f'N{current_row}'] = prop['vintage']
        if prop.get('avg_unit_size'):
            sheet[f'O{current_row}'] = prop['avg_unit_size']

        # Subject rent data
        if rent_comps.get('subject_current_rent_psf'):
            sheet[f'Q{current_row}'] = rent_comps['subject_current_rent_psf']
        if rent_comps.get('subject_current_rent'):
            sheet[f'R{current_row}'] = rent_comps['subject_current_rent']

        # Subject vacancy -> occupancy
        if prop.get('vacancy_rate'):
            sheet[f'W{current_row}'] = (100 - prop['vacancy_rate']) / 100

        # === SUBJECT PROPERTY - Unit Mix Data (Rent by type, unit counts, concessions) ===
        unit_mix = prop.get('unit_mix', [])
        if unit_mix:
            # Create lookup by bedroom count
            unit_by_beds = {u.get('bedrooms'): u for u in unit_mix}

            # S: Studio rent (bedrooms=0)
            if 0 in unit_by_beds and unit_by_beds[0].get('asking_rent_per_unit'):
                sheet[f'S{current_row}'] = unit_by_beds[0]['asking_rent_per_unit']

            # T: 1 Bed rent (bedrooms=1)
            if 1 in unit_by_beds and unit_by_beds[1].get('asking_rent_per_unit'):
                sheet[f'T{current_row}'] = unit_by_beds[1]['asking_rent_per_unit']

            # U: 2 Bed rent (bedrooms=2)
            if 2 in unit_by_beds and unit_by_beds[2].get('asking_rent_per_unit'):
                sheet[f'U{current_row}'] = unit_by_beds[2]['asking_rent_per_unit']

            # V: 3 Bed rent (bedrooms=3)
            if 3 in unit_by_beds and unit_by_beds[3].get('asking_rent_per_unit'):
                sheet[f'V{current_row}'] = unit_by_beds[3]['asking_rent_per_unit']

            # X: Concession % (use average or first available)
            concessions = [u.get('concessions_pct') for u in unit_mix if u.get('concessions_pct') is not None]
            if concessions:
                avg_concession = sum(concessions) / len(concessions)
                sheet[f'X{current_row}'] = avg_concession / 100

            # Y: #Studio units (bedrooms=0)
            if 0 in unit_by_beds and unit_by_beds[0].get('units'):
                sheet[f'Y{current_row}'] = unit_by_beds[0]['units']

            # Z: #1 Bed units (bedrooms=1)
            if 1 in unit_by_beds and unit_by_beds[1].get('units'):
                sheet[f'Z{current_row}'] = unit_by_beds[1]['units']

            # AA: #2 Bed units (bedrooms=2)
            if 2 in unit_by_beds and unit_by_beds[2].get('units'):
                sheet[f'AA{current_row}'] = unit_by_beds[2]['units']

            # AB: #3 Bed units (bedrooms=3)
            if 3 in unit_by_beds and unit_by_beds[3].get('units'):
                sheet[f'AB{current_row}'] = unit_by_beds[3]['units']

        # AC: Neighborhood (from market data if available)
        market = extracted_data.get('market', {})
        if market.get('property_submarket'):
            sheet[f'AC{current_row}'] = market['property_submarket']

        current_row += 1

        # === COMPARABLE PROPERTIES (Row 10+) ===
        for i, comp in enumerate(comps[:17]):  # Limit to 17 comps
            row = current_row + i

            # D: Building Name
            if comp.get('name'):
                sheet[f'D{row}'] = comp['name']

            # E: Address
            if comp.get('address'):
                sheet[f'E{row}'] = comp['address']

            # F: City
            if comp.get('city'):
                sheet[f'F{row}'] = comp['city']

            # G: State
            if comp.get('state'):
                sheet[f'G{row}'] = comp['state']

            # L: Units
            if comp.get('units'):
                sheet[f'L{row}'] = comp['units']

            # M: Stories
            if comp.get('stories'):
                sheet[f'M{row}'] = comp['stories']

            # N: Year Built/Renovated
            if comp.get('year_built_display'):
                sheet[f'N{row}'] = comp['year_built_display']
            elif comp.get('year_built'):
                sheet[f'N{row}'] = comp['year_built']

            # O: Avg SF
            if comp.get('avg_sf'):
                sheet[f'O{row}'] = comp['avg_sf']

            # P: Distance (mi away)
            if comp.get('distance'):
                sheet[f'P{row}'] = comp['distance']

            # Q: Rent/SF
            if comp.get('rent_psf'):
                sheet[f'Q{row}'] = comp['rent_psf']

            # R: Rent/Unit (calculate if not available)
            if comp.get('rent_per_unit'):
                sheet[f'R{row}'] = comp['rent_per_unit']
            elif comp.get('rent_psf') and comp.get('avg_sf'):
                sheet[f'R{row}'] = round(comp['rent_psf'] * comp['avg_sf'])

            # S: Studio rent
            if comp.get('rent_studio'):
                sheet[f'S{row}'] = comp['rent_studio']

            # T: 1 Bed rent
            if comp.get('rent_1bed'):
                sheet[f'T{row}'] = comp['rent_1bed']

            # U: 2 Bed rent
            if comp.get('rent_2bed'):
                sheet[f'U{row}'] = comp['rent_2bed']

            # V: 3 Bed rent
            if comp.get('rent_3bed'):
                sheet[f'V{row}'] = comp['rent_3bed']

            # W: Occupancy % (convert from vacancy)
            if comp.get('vacancy'):
                sheet[f'W{row}'] = (100 - comp['vacancy']) / 100
            elif comp.get('occupancy'):
                sheet[f'W{row}'] = comp['occupancy'] / 100

            # X: Concession %
            if comp.get('concession_pct'):
                sheet[f'X{row}'] = comp['concession_pct'] / 100
            elif comp.get('concession') is not None:
                sheet[f'X{row}'] = comp['concession'] / 100

            # Y: #Studio (unit count)
            if comp.get('unit_count_studio'):
                sheet[f'Y{row}'] = comp['unit_count_studio']

            # Z: #1 Beds (unit count)
            if comp.get('unit_count_1bed'):
                sheet[f'Z{row}'] = comp['unit_count_1bed']

            # AA: #2 Beds (unit count)
            if comp.get('unit_count_2bed'):
                sheet[f'AA{row}'] = comp['unit_count_2bed']

            # AB: #3 Beds (unit count)
            if comp.get('unit_count_3bed'):
                sheet[f'AB{row}'] = comp['unit_count_3bed']

            # AC: Neighborhood
            if comp.get('neighborhood'):
                sheet[f'AC{row}'] = comp['neighborhood']

        # Apply center alignment to all data cells (rows 9-26, columns B-AC)
        center_align = Alignment(horizontal='center', vertical='center')
        rent_comp_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                         'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']
        for row in range(9, 27):  # Rows 9-26
            for col in rent_comp_cols:
                sheet[f'{col}{row}'].alignment = center_align

        print(f"[OK] Wrote subject + {len(comps[:17])} rent comparables to Rent Comps sheet")

    def write_sale_comps(self, extracted_data: Dict[str, Any], config: Dict[str, Any] = None):
        """
        Write sale comparables to the Sale Comps sheet.
        Subject property goes first, then comps.

        Excel columns (Row 7 headers, data starts row 8):
        B: Address, C: Name, D: Rating, E: Yr Blt/Renov, F: Type,
        G: Units, H: Dist (mi), I: Sale Date, J: Sale Price,
        K: Price/Unit, L: Price/SF, M: Cap Rate, N: Submarket
        """
        if not self.workbook:
            return

        if 'Sale Comps' not in self.workbook.sheetnames:
            print("[INFO] No 'Sale Comps' sheet found in workbook - skipping sale comps")
            return

        sheet = self.workbook['Sale Comps']
        sale_comps = extracted_data.get('sale_comps', {})
        comps = sale_comps.get('comparable_properties', [])
        prop = extracted_data.get('property', {})

        # Data starts at row 8 (row 7 is headers)
        start_row = 8
        current_row = start_row

        # === SUBJECT PROPERTY FIRST (Row 8) ===
        # Get address info from config
        if config and config.get('property_details'):
            details = config['property_details']
            if details.get('address'):
                sheet[f'B{current_row}'] = details['address']

        # Get property name from config
        if config and config.get('property_name'):
            sheet[f'C{current_row}'] = config['property_name']

        # Subject property data
        if prop.get('vintage'):
            sheet[f'E{current_row}'] = prop['vintage']
        if prop.get('units'):
            sheet[f'G{current_row}'] = prop['units']

        current_row += 1

        # === COMPARABLE PROPERTIES (Row 9+) ===
        if not comps:
            print("[INFO] No sale comps data to write")
            return

        for i, comp in enumerate(comps[:15]):  # Limit to 15 sale comps
            row = current_row + i

            # B: Address
            if comp.get('address'):
                sheet[f'B{row}'] = comp['address']

            # C: Name
            if comp.get('name'):
                sheet[f'C{row}'] = comp['name']

            # E: Year Built/Renovated
            if comp.get('year_built'):
                sheet[f'E{row}'] = comp['year_built']

            # F: Type
            if comp.get('type'):
                sheet[f'F{row}'] = comp['type']

            # G: Units
            if comp.get('units'):
                sheet[f'G{row}'] = comp['units']

            # H: Distance (mi)
            if comp.get('distance'):
                sheet[f'H{row}'] = comp['distance']

            # I: Sale Date
            if comp.get('sale_date'):
                sheet[f'I{row}'] = comp['sale_date']

            # J: Sale Price
            if comp.get('sale_price'):
                sheet[f'J{row}'] = comp['sale_price']

            # K: Price Per Unit
            if comp.get('price_per_unit'):
                sheet[f'K{row}'] = comp['price_per_unit']

            # L: Price Per SF
            if comp.get('price_per_sf'):
                sheet[f'L{row}'] = comp['price_per_sf']

            # M: Cap Rate
            if comp.get('cap_rate'):
                sheet[f'M{row}'] = comp['cap_rate'] / 100

            # N: Submarket
            if comp.get('submarket'):
                sheet[f'N{row}'] = comp['submarket']

        # Apply formatting to all data cells (rows 8-22)
        center_align = Alignment(horizontal='center', vertical='center')
        sale_comp_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        for row in range(8, 23):  # Rows 8-22
            for col in sale_comp_cols:
                sheet[f'{col}{row}'].alignment = center_align
            # Apply currency format to J (Sale Price), K (Price/Unit), L (Price/SF)
            sheet[f'J{row}'].number_format = '"$"#,##0'
            sheet[f'K{row}'].number_format = '"$"#,##0'
            sheet[f'L{row}'].number_format = '"$"#,##0'
            # Apply percentage format to M (Cap Rate)
            sheet[f'M{row}'].number_format = '0.0%'

        print(f"[OK] Wrote subject + {len(comps[:15])} sale comparables to Sale Comps sheet")

    def insert_map_image(self, sheet_name: str, cell: str, image_path: str,
                         width_inches: float = 5.0, height_inches: float = 3.5) -> bool:
        """
        Insert a map screenshot image into a specific sheet and cell.

        Args:
            sheet_name: Name of the Excel sheet
            cell: Cell reference (e.g., 'B39') - top-left anchor
            image_path: Path to the PNG image file
            width_inches: Desired width in inches
            height_inches: Desired height in inches

        Returns:
            True if successful, False otherwise
        """
        import os
        try:
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image
        except ImportError:
            print("PIL not installed. Skipping image insertion.")
            return False

        if not self.workbook:
            print("No workbook loaded")
            return False

        if not os.path.exists(image_path):
            print(f"Image not found: {image_path}")
            return False

        if sheet_name not in self.workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Available: {self.workbook.sheetnames}")
            return False

        ws = self.workbook[sheet_name]

        try:
            # Open and resize image preserving aspect ratio
            img = Image.open(image_path)
            orig_width, orig_height = img.size

            # Convert to RGB if needed (for PNG with transparency)
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')

            # Resize image preserving aspect ratio (96 DPI is Excel's default)
            target_width = int(width_inches * 96)
            aspect_ratio = orig_height / orig_width
            target_height = int(target_width * aspect_ratio)
            img = img.resize((target_width, target_height), Image.LANCZOS)

            # Save resized image to temp directory
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            temp_path = temp_file.name
            temp_file.close()
            img.save(temp_path, 'PNG')

            # Create Excel image object
            xl_img = XLImage(temp_path)
            xl_img.anchor = cell

            # Add image to worksheet
            ws.add_image(xl_img)

            print(f"[OK] Inserted map image at {sheet_name}!{cell}")

            # Clean up temp file after saving (done in save method)
            self._temp_image_files = getattr(self, '_temp_image_files', [])
            self._temp_image_files.append(temp_path)

            return True

        except Exception as e:
            print(f"Error inserting image: {e}")
            return False

    def insert_cover_maps(self, location_png: str, parcel_png: str,
                          location_html: str, parcel_html: str) -> bool:
        """
        Insert location and parcel maps into Screener Cover sheet with aligned heights.

        Args:
            location_png: Path to location map screenshot
            parcel_png: Path to parcel/satellite screenshot
            location_html: Path to interactive location HTML map
            parcel_html: Path to interactive parcel HTML map

        Returns:
            True if successful, False otherwise
        """
        import os
        try:
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image
        except ImportError:
            print("PIL not installed. Skipping map insertion.")
            return False

        if not self.workbook:
            print("No workbook loaded")
            return False

        sheet_name = 'Screener Cover'
        if sheet_name not in self.workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found")
            return False

        ws = self.workbook[sheet_name]

        try:
            # Get parcel dimensions to calculate aligned height
            parcel_img = Image.open(parcel_png)
            parcel_w, parcel_h = parcel_img.size
            parcel_display_width = 4.0 * 96  # 4 inches at 96 DPI
            parcel_aspect = parcel_h / parcel_w
            parcel_display_height = parcel_display_width * parcel_aspect

            # Calculate location map size to match parcel height
            location_img = Image.open(location_png)
            loc_w, loc_h = location_img.size
            loc_aspect = loc_h / loc_w
            # Target height = parcel height, calculate width from aspect ratio
            target_height = parcel_display_height
            target_width = target_height / loc_aspect

            # Process and insert location map at B39
            import tempfile
            if location_img.mode in ('RGBA', 'P'):
                location_img = location_img.convert('RGB')
            location_img = location_img.resize((int(target_width), int(target_height)), Image.LANCZOS)
            loc_temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            loc_temp = loc_temp_file.name
            loc_temp_file.close()
            location_img.save(loc_temp, 'PNG')
            xl_loc = XLImage(loc_temp)
            xl_loc.anchor = 'B39'
            ws.add_image(xl_loc)
            self._temp_image_files = getattr(self, '_temp_image_files', [])
            self._temp_image_files.append(loc_temp)

            # Process and insert parcel map at E39
            if parcel_img.mode in ('RGBA', 'P'):
                parcel_img = parcel_img.convert('RGB')
            parcel_img = parcel_img.resize((int(parcel_display_width), int(parcel_display_height)), Image.LANCZOS)
            parcel_temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            parcel_temp = parcel_temp_file.name
            parcel_temp_file.close()
            parcel_img.save(parcel_temp, 'PNG')
            xl_parcel = XLImage(parcel_temp)
            xl_parcel.anchor = 'E39'
            ws.add_image(xl_parcel)
            self._temp_image_files.append(parcel_temp)

            # Add hyperlinks at row 56
            ws['B56'] = 'Open Interactive Map'
            ws['B56'].hyperlink = location_html
            ws['B56'].style = 'Hyperlink'

            ws['E56'] = 'Open Satellite View'
            ws['E56'].hyperlink = parcel_html
            ws['E56'].style = 'Hyperlink'

            # Track temp files for cleanup
            self._temp_image_files = getattr(self, '_temp_image_files', [])
            self._temp_image_files.extend([loc_temp, parcel_temp])

            print(f"[OK] Inserted cover maps (location: {target_width/96:.1f}\" x {target_height/96:.1f}\", parcel: 4.0\" x {parcel_display_height/96:.1f}\")")
            return True

        except Exception as e:
            print(f"Error inserting cover maps: {e}")
            return False

    def insert_stage1_income_map(self, income_png: str, income_html: str = '') -> bool:
        """
        Insert income map into Stage 1 sheet at C12.

        Args:
            income_png: Path to income map screenshot
            income_html: Path to interactive income HTML map (optional, for hyperlink)

        Returns:
            True if successful, False otherwise
        """
        import os
        try:
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image
        except ImportError:
            print("PIL not installed. Skipping income map insertion.")
            return False

        if not self.workbook:
            print("No workbook loaded")
            return False

        sheet_name = 'Stage 1'
        if sheet_name not in self.workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found")
            return False

        ws = self.workbook[sheet_name]

        try:
            # Open and process income map - preserve aspect ratio
            income_img = Image.open(income_png)
            img_w, img_h = income_img.size

            # Target width 11 inches, calculate height to preserve aspect ratio
            target_width = 11.0 * 96  # 11 inches at 96 DPI
            aspect = img_h / img_w
            target_height = target_width * aspect  # Preserves aspect ratio

            # Convert and resize (no stretching)
            import tempfile
            if income_img.mode in ('RGBA', 'P'):
                income_img = income_img.convert('RGB')
            income_img = income_img.resize((int(target_width), int(target_height)), Image.LANCZOS)

            # Save temp file for Excel
            income_temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            income_temp = income_temp_file.name
            income_temp_file.close()
            income_img.save(income_temp, 'PNG')

            # Insert image at C12
            xl_income = XLImage(income_temp)
            xl_income.anchor = 'C12'
            ws.add_image(xl_income)
            self._temp_image_files = getattr(self, '_temp_image_files', [])
            self._temp_image_files.append(income_temp)

            # Add hyperlink at row 23
            if income_html and os.path.exists(income_html):
                ws['C23'] = 'Open Interactive Income Map'
                ws['C23'].hyperlink = income_html
                ws['C23'].style = 'Hyperlink'

            # Track temp files for cleanup
            self._temp_image_files = getattr(self, '_temp_image_files', [])
            self._temp_image_files.append(income_temp)

            print(f"[OK] Inserted income map on Stage 1 at C12 ({target_width/96:.1f}\" x {target_height/96:.1f}\")")
            return True

        except Exception as e:
            print(f"Error inserting income map on Stage 1: {e}")
            return False

    def insert_stage1_flood_map(self, flood_map_path: str) -> bool:
        """
        Insert flood zone map screenshot on Stage 1 sheet at cell C51.
        Same dimensions as income map (11" wide).

        Args:
            flood_map_path: Path to the flood map PNG screenshot

        Returns:
            True if successful, False otherwise
        """
        if not self.workbook:
            print("Error: No workbook loaded")
            return False

        if not os.path.exists(flood_map_path):
            print(f"Error: Flood map file not found: {flood_map_path}")
            return False

        try:
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image as PILImage

            # Get or create Stage 1 sheet
            if 'Stage 1' not in self.workbook.sheetnames:
                print("Warning: 'Stage 1' sheet not found")
                return False

            ws = self.workbook['Stage 1']

            # Load the image
            pil_img = PILImage.open(flood_map_path)
            orig_width, orig_height = pil_img.size

            # Target width: 11 inches (same as income map) = 1056 pixels at 96 DPI
            target_width = 11.0 * 96
            scale = target_width / orig_width
            target_height = orig_height * scale

            # Create temp resized image
            import tempfile
            flood_temp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            flood_temp_path = flood_temp.name
            flood_temp.close()

            resized = pil_img.resize((int(target_width), int(target_height)), PILImage.LANCZOS)
            resized.save(flood_temp_path, 'PNG')
            pil_img.close()

            # Insert into Excel at C51
            xl_img = XLImage(flood_temp_path)
            xl_img.anchor = 'C51'
            ws.add_image(xl_img)

            # Add hyperlink to HTML file at C62 (below the map)
            html_path = flood_map_path.replace('.png', '.html')
            if os.path.exists(html_path):
                ws['C62'] = 'Click to open interactive flood map'
                ws['C62'].hyperlink = html_path
                ws['C62'].style = 'Hyperlink'

            # Track temp files for cleanup
            self._temp_image_files = getattr(self, '_temp_image_files', [])
            self._temp_image_files.append(flood_temp_path)

            print(f"[OK] Inserted flood map on Stage 1 at C51 ({target_width/96:.1f}\" x {target_height/96:.1f}\")")
            return True

        except Exception as e:
            print(f"Error inserting flood map on Stage 1: {e}")
            return False

    def write_flood_zone_data(self, flood_info: dict) -> bool:
        """
        Write flood zone information to Stage 1 sheet above the flood map.

        Args:
            flood_info: Dict with keys 'zone', 'flood_risk', 'zone_subtype'

        Returns:
            True if successful, False otherwise
        """
        if not self.workbook:
            print("Error: No workbook loaded")
            return False

        if not flood_info:
            print("Warning: No flood zone info provided")
            return False

        try:
            # Write to Stage 1 sheet
            if 'Stage 1' not in self.workbook.sheetnames:
                print("Warning: 'Stage 1' sheet not found")
                return False

            ws = self.workbook['Stage 1']

            # Write flood zone info at C49 and C50 (above the flood map at C51)
            zone = flood_info.get('zone', 'Unknown')
            risk = flood_info.get('flood_risk', 'Unknown')

            ws['C49'] = f"FEMA Flood Zone: {zone}"
            ws['C50'] = f"Risk Level: {risk}"

            print(f"[OK] Wrote flood zone data: Zone {zone} - {risk}")
            return True

        except Exception as e:
            print(f"Error writing flood zone data: {e}")
            return False

    def save(self):
        """Save the workbook and clean up temp files."""
        if not self.workbook:
            raise ValueError("No workbook to save.")

        self.workbook.save(self.output_file)
        print(f"\n[OK] Saved updated screener to: {self.output_file}")

        # Clean up temp image files
        temp_files = getattr(self, '_temp_image_files', [])
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass  # Ignore cleanup errors

    def get_changes_summary(self) -> str:
        """Get summary of all changes made."""
        if not self.changes_made:
            return "No data was populated."

        summary = [f"\n=== DATA INPUTS POPULATED ({len(self.changes_made)} fields) ===\n"]

        # Group by section
        for change in self.changes_made:
            summary.append(
                f"  {change['field']}: {change['value']} ({change['source']})"
            )

        return "\n".join(summary)

    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()


if __name__ == "__main__":
    # Test the writer
    print("Excel Writer module loaded successfully")
