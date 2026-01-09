"""
Build transparent Scoring Reference sheet V2 - Improved layout
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load template
wb = openpyxl.load_workbook(r'C:\Users\carso\OneDrive - UCB-O365\Work\RMP\Screener\RMP Screener_PreLinked_v3.xlsx')

# Delete and recreate Scoring Reference sheet
if 'Scoring Reference' in wb.sheetnames:
    del wb['Scoring Reference']
ws = wb.create_sheet('Scoring Reference')

# Styles - Dark green header like other sheets
header_fill = PatternFill(start_color='375623', end_color='375623', fill_type='solid')  # Dark green
header_font = Font(bold=True, size=11, color='FFFFFF')
subheader_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')  # Light green
subheader_font = Font(bold=True, size=11)
scale_header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Very light green
final_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow
method_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Light gray
normal_font = Font(size=11)
bold_font = Font(bold=True, size=11)
wrap_align = Alignment(wrap_text=True, vertical='top')

# Set column widths
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 3   # Spacer
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 50  # Methodology

# Helper for XLOOKUP formulas
def xlookup(field):
    return f'=_xlfn.XLOOKUP("{field}",\'Data Inputs\'!$B:$B,\'Data Inputs\'!$C:$C)'

def add_section_header(row, text):
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = text
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    return row + 1

def add_column_headers(row):
    headers = ['Field', 'Value', 'Python Check', '', 'Scale', 'Score', 'Methodology']
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = subheader_font
        cell.fill = subheader_fill
    return row + 1

def add_methodology(row, text, num_rows=3):
    """Add merged methodology explanation block"""
    ws.merge_cells(f'G{row}:G{row+num_rows-1}')
    ws[f'G{row}'] = text
    ws[f'G{row}'].font = normal_font
    ws[f'G{row}'].alignment = wrap_align
    ws[f'G{row}'].fill = method_fill

# Title
ws.merge_cells('A1:G1')
ws['A1'] = 'STAGE 2 SCORING CALCULATOR'
ws['A1'].font = Font(bold=True, size=14)
ws['A2'] = 'All scores calculated via Excel formulas | Column C shows Python value for verification'
ws['A2'].font = Font(size=11, italic=True)

row = 4

# Track final score rows for Stage 2 reference
final_score_rows = {}

#===============================================================================
# SUPPLY-DEMAND DRIVERS (5% weight)
#===============================================================================
row = add_section_header(row, 'SUPPLY-DEMAND DRIVERS (5% weight)')
row = add_column_headers(row)

sd_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Absorption (12mo)', xlookup('12 Mo Absorption Units')
ws[f'C{row}'] = xlookup('SD: Absorption (12mo)')
ws[f'E{row}'], ws[f'F{row}'] = 'Ratio >=2.0', 10
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Delivered (12mo)', xlookup('12 Mo Delivered Units')
ws[f'C{row}'] = xlookup('SD: Delivered (12mo)')
ws[f'E{row}'], ws[f'F{row}'] = '1.5 - 2.0', 8
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Under Construction', xlookup('Under Construction Units')
ws[f'C{row}'] = xlookup('SD: Under Construction')
ws[f'E{row}'], ws[f'F{row}'] = '1.0 - 1.5', 6
row += 1

ws[f'A{row}'] = 'Absorption Ratio'
ws[f'B{row}'] = f'=IF(B{sd_start+1}=0,"N/A",ROUND(B{sd_start}/B{sd_start+1},2))'
ws[f'C{row}'] = xlookup('SD: Absorption Ratio')
ws[f'E{row}'], ws[f'F{row}'] = '0.5 - 1.0', 4
row += 1

ws[f'A{row}'] = 'Pipeline Ratio'
ws[f'B{row}'] = f'=IF(B{sd_start}<=0,"N/A",ROUND(B{sd_start+2}/B{sd_start},2))'
ws[f'C{row}'] = xlookup('SD: Pipeline Ratio')
ws[f'E{row}'], ws[f'F{row}'] = '<0.5', 2
row += 1

ws[f'A{row}'] = 'Base Score'
ws[f'B{row}'] = f'=IF(B{sd_start}<0,3,IF(B{sd_start+1}=0,IF(B{sd_start}>0,9,5),IF(B{sd_start+3}>=2,10,IF(B{sd_start+3}>=1.5,8,IF(B{sd_start+3}>=1,6,IF(B{sd_start+3}>=0.5,4,2))))))'
ws[f'C{row}'] = xlookup('SD: Base Score')
ws[f'E{row}'] = 'Pipeline Adj'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'] = 'Pipeline Adjustment'
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(B{sd_start+4})),0,IF(B{sd_start+4}>1.5,-2,IF(B{sd_start+4}>1,-1,IF(B{sd_start+4}<0.5,1,0))))'
ws[f'C{row}'] = xlookup('SD: Pipeline Adjustment')
ws[f'E{row}'], ws[f'F{row}'] = '>1.5x absorb', -2
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=MAX(1,MIN(10,B{row-2}+B{row-1}))'
ws[f'B{row}'].fill = final_fill
ws[f'C{row}'] = xlookup('SD: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '>1.0x absorb', -1
final_score_rows['SD'] = row
row += 1

ws[f'E{row}'], ws[f'F{row}'] = '<0.5x absorb', '+1'
row += 1

# Methodology
method_start = sd_start
add_methodology(method_start,
    "LOGIC: Measures if submarket absorbs new supply.\n\n"
    "1. Absorption Ratio = Units Absorbed / Units Delivered\n"
    "   - >2.0 means demand outpaces supply (strong)\n"
    "   - <1.0 means oversupply risk\n\n"
    "2. Pipeline Adjustment = future supply pressure\n"
    "   - Heavy pipeline (>1.5x absorption) = -2 penalty\n"
    "   - Light pipeline (<0.5x) = +1 bonus\n\n"
    "Final = Base + Pipeline Adj (capped 1-10)", 9)

row += 2

#===============================================================================
# SUBMARKET SUPPLY-DEMAND OUTLOOK (10% weight)
#===============================================================================
row = add_section_header(row, 'SUBMARKET SUPPLY-DEMAND OUTLOOK (10% weight)')
row = add_column_headers(row)

so_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Current Vacancy %', xlookup('Submarket Vacancy Rate %')
ws[f'C{row}'] = xlookup('SO: Current Vacancy %')
ws[f'E{row}'] = 'Level Adj'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Vacancy YoY Change', xlookup('Vacancy YoY Change %')
ws[f'C{row}'] = xlookup('SO: Vacancy YoY Change')
ws[f'E{row}'], ws[f'F{row}'] = '>2ppt below avg', '+2'
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Historical Avg Vacancy', xlookup('Vacancy Historical Avg %')
ws[f'C{row}'] = xlookup('SO: Historical Avg Vacancy')
ws[f'E{row}'], ws[f'F{row}'] = '0.5-2ppt below', '+1'
row += 1

ws[f'A{row}'] = 'Vacancy vs Historical'
ws[f'B{row}'] = f'=IF(OR(B{so_start}="",B{so_start+2}=""),"N/A",ROUND(B{so_start}-B{so_start+2},1))'
ws[f'C{row}'] = xlookup('SO: Vacancy vs Historical')
ws[f'E{row}'], ws[f'F{row}'] = 'Within 0.5ppt', '0'
row += 1

ws[f'A{row}'] = 'Level Adjustment'
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(B{row-1})),0,IF(B{row-1}<=-2,2,IF(B{row-1}<=-0.5,1,IF(B{row-1}<=0.5,0,IF(B{row-1}<=2,-1,-2)))))'
ws[f'C{row}'] = xlookup('SO: Level Adjustment')
ws[f'E{row}'], ws[f'F{row}'] = '0.5-2ppt above', '-1'
row += 1

ws[f'A{row}'] = 'Trend Adjustment'
ws[f'B{row}'] = f'=IF(B{so_start+1}="",0,IF(B{so_start+1}<=-1,3,IF(B{so_start+1}<=-0.5,2,IF(B{so_start+1}<0,1,IF(B{so_start+1}=0,0,IF(B{so_start+1}<=0.5,-1,IF(B{so_start+1}<=1,-2,-3)))))))'
ws[f'C{row}'] = xlookup('SO: Trend Adjustment')
ws[f'E{row}'], ws[f'F{row}'] = '>2ppt above', '-2'
row += 1

ws[f'A{row}'] = 'Pipeline Adjustment'
ws[f'B{row}'] = f'=LET(pr,B{so_start+8},IF(NOT(ISNUMBER(pr)),0,IF(pr<0.5,1,IF(pr<=1,0,IF(pr<=1.5,-1,-2)))))'
ws[f'C{row}'] = xlookup('SO: Pipeline Adjustment')
ws[f'E{row}'] = 'Trend Adj'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'] = 'Pipeline Ratio'
ws[f'B{row}'] = f'=IF(B{sd_start}<=0,"N/A",ROUND(B{sd_start+2}/B{sd_start},2))'
ws[f'C{row}'] = xlookup('SO: Pipeline Ratio')
ws[f'E{row}'], ws[f'F{row}'] = 'YoY <-1%', '+3'
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=MAX(1,MIN(10,5+B{row-5}+B{row-4}+B{row-3}))'
ws[f'B{row}'].fill = final_fill
ws[f'C{row}'] = xlookup('SO: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = 'YoY -0.5 to -1%', '+2'
final_score_rows['SO'] = row
row += 1

ws[f'E{row}'], ws[f'F{row}'] = 'YoY 0 to -0.5%', '+1'
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'YoY 0 to +0.5%', '-1'
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'YoY +0.5 to +1%', '-2'
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'YoY >+1%', '-3'
row += 1

add_methodology(so_start,
    "LOGIC: Is the submarket getting tighter or looser?\n\n"
    "Base Score = 5 (neutral)\n\n"
    "1. Level Adj: Current vs Historical vacancy\n"
    "   - Below avg = positive (tighter market)\n"
    "   - Above avg = negative (looser market)\n\n"
    "2. Trend Adj: YoY vacancy change (most important)\n"
    "   - Falling vacancy = positive\n"
    "   - Rising vacancy = negative\n\n"
    "3. Pipeline Adj: Future supply pressure\n\n"
    "Final = 5 + Level + Trend + Pipeline (capped 1-10)", 13)

row += 2

#===============================================================================
# MIGRATION / GDP GROWTH (3% weight)
#===============================================================================
row = add_section_header(row, 'MIGRATION / GDP GROWTH (3% weight)')
row = add_column_headers(row)

mg_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Emp Growth - Market', xlookup('Employment Growth - Market')
ws[f'C{row}'] = xlookup('MG: Employment Growth - Market')
ws[f'E{row}'] = 'Emp vs US'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Emp Growth - US', xlookup('Employment Growth - US')
ws[f'C{row}'] = xlookup('MG: Employment Growth - US')
ws[f'E{row}'], ws[f'F{row}'] = '>=+1%', 10
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Pop Growth (5mi) %', xlookup('Population Growth % (5mi)')
ws[f'C{row}'] = xlookup('MG: Pop Growth (5mi)')
ws[f'E{row}'], ws[f'F{row}'] = '+0.5 to +1%', 8
row += 1

ws[f'A{row}'] = 'Employment vs US'
ws[f'B{row}'] = f'=IF(OR(B{mg_start}="",B{mg_start+1}=""),"N/A",ROUND(B{mg_start}-B{mg_start+1},2))'
ws[f'C{row}'] = xlookup('MG: Employment vs US')
ws[f'E{row}'], ws[f'F{row}'] = '0 to +0.5%', 6
row += 1

ws[f'A{row}'] = 'Employment Score'
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(B{row-1})),"",IF(B{row-1}>=1,10,IF(B{row-1}>=0.5,8,IF(B{row-1}>=0,6,IF(B{row-1}>=-0.5,5,IF(B{row-1}>=-1,4,IF(B{row-1}>=-1.5,3,2)))))))'
ws[f'C{row}'] = xlookup('MG: Employment Score')
ws[f'E{row}'], ws[f'F{row}'] = '-0.5 to 0%', 5
row += 1

ws[f'A{row}'] = 'Population Score'
ws[f'B{row}'] = f'=IF(B{mg_start+2}="","",IF(B{mg_start+2}>=10,10,IF(B{mg_start+2}>=8,9,IF(B{mg_start+2}>=6,8,IF(B{mg_start+2}>=4,7,IF(B{mg_start+2}>=2,6,IF(B{mg_start+2}>=0,5,IF(B{mg_start+2}>=-2,4,2))))))))'
ws[f'C{row}'] = xlookup('MG: Population Score')
ws[f'E{row}'], ws[f'F{row}'] = '<-1.5%', 2
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(AND(B{row-2}="",B{row-1}=""),"",ROUND(AVERAGE(B{row-2},B{row-1}),0))'
ws[f'B{row}'].fill = final_fill
ws[f'C{row}'] = xlookup('MG: Final Score')
final_score_rows['MG'] = row
row += 1

add_methodology(mg_start,
    "LOGIC: Is the local economy growing?\n\n"
    "Two sub-scores averaged:\n\n"
    "1. Employment Score\n"
    "   - Market job growth vs US average\n"
    "   - Outperforming US = higher score\n\n"
    "2. Population Score\n"
    "   - 5-year population growth projection\n"
    "   - Proxy for migration trends\n\n"
    "Final = Average of both scores", 7)

row += 2

#===============================================================================
# PARKING RATIO (3% weight)
#===============================================================================
row = add_section_header(row, 'PARKING RATIO (3% weight)')
row = add_column_headers(row)

pr_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Parking Ratio', xlookup('PR: Parking Ratio')
ws[f'C{row}'] = xlookup('PR: Parking Ratio')
ws[f'E{row}'] = 'Ratio'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Surface Spaces', xlookup('PR: Surface Spaces')
ws[f'C{row}'] = xlookup('PR: Surface Spaces')
ws[f'E{row}'], ws[f'F{row}'] = '>=2.0', 10
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Covered Spaces', xlookup('PR: Covered Spaces')
ws[f'C{row}'] = xlookup('PR: Covered Spaces')
ws[f'E{row}'], ws[f'F{row}'] = '1.5 - 2.0', 9
row += 1

ws[f'A{row}'] = 'Base Score'
ws[f'B{row}'] = f'=IF(B{pr_start}="","",IF(B{pr_start}>=2,10,IF(B{pr_start}>=1.5,9,IF(B{pr_start}>=1.25,8,IF(B{pr_start}>=1,7,IF(B{pr_start}>=0.75,5,IF(B{pr_start}>=0.5,3,2)))))))'
ws[f'C{row}'] = xlookup('PR: Base Score')
ws[f'E{row}'], ws[f'F{row}'] = '1.25 - 1.5', 8
row += 1

ws[f'A{row}'] = 'Underground Penalty'
ws[f'B{row}'] = f'=IF(OR(B{pr_start+1}="",B{pr_start+2}=""),0,IF(AND(B{pr_start+1}=0,B{pr_start+2}>0),-1,0))'
ws[f'C{row}'] = xlookup('PR: Underground Penalty')
ws[f'E{row}'], ws[f'F{row}'] = '1.0 - 1.25', 7
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(B{row-2}="","",MAX(1,MIN(10,B{row-2}+B{row-1})))'
ws[f'B{row}'].fill = final_fill
ws[f'C{row}'] = xlookup('PR: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '0.75 - 1.0', 5
final_score_rows['PR'] = row
row += 1

ws[f'E{row}'], ws[f'F{row}'] = '0.5 - 0.75', 3
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '<0.5', 2
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'Underground-only', '-1'
row += 1

add_methodology(pr_start,
    "LOGIC: Adequate parking for suburban workforce housing.\n\n"
    "Target: 1:1 or better ratio\n\n"
    "Penalty: -1 if underground/garage only (no surface)\n"
    "- Higher maintenance costs\n"
    "- Less convenient for residents\n\n"
    "Final = Base Score + Penalty (capped 1-10)", 9)

row += 2

#===============================================================================
# AMENITIES & LIFESTYLE (5% weight)
#===============================================================================
row = add_section_header(row, 'AMENITIES & LIFESTYLE (5% weight)')
row = add_column_headers(row)

am_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Site Score (max 5)', xlookup('AM: Site Score')
ws[f'C{row}'] = xlookup('AM: Site Score')
ws[f'E{row}'] = 'Site Amenities'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Unit Score (max 5)', xlookup('AM: Unit Score')
ws[f'C{row}'] = xlookup('AM: Unit Score')
ws[f'E{row}'], ws[f'F{row}'] = 'Pool', 1.0
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(AND(B{am_start}="",B{am_start+1}=""),"",MAX(1,MIN(10,B{am_start}+B{am_start+1})))'
ws[f'B{row}'].fill = final_fill
ws[f'C{row}'] = xlookup('AM: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = 'Fitness Center', 1.0
final_score_rows['AM'] = row
row += 1

ws[f'E{row}'], ws[f'F{row}'] = 'Clubhouse', 1.0
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'Dog Park', 0.5
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'Playground', 0.5
row += 1
ws[f'E{row}'] = 'Unit Amenities'
ws[f'E{row}'].font = bold_font
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'In-Unit W/D', 1.5
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'W/D Hookup', 0.75
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'A/C', 1.0
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'Dishwasher', 0.5
row += 1

add_methodology(am_start,
    "LOGIC: Does property meet market expectations?\n\n"
    "Site Score (max 5 pts):\n"
    "- Pool, Fitness, Clubhouse = 1.0 each\n"
    "- Dog Park, Playground = 0.5 each\n\n"
    "Unit Score (max 5 pts):\n"
    "- In-Unit W/D = 1.5\n"
    "- A/C = 1.0\n"
    "- Dishwasher, Balcony = 0.5 each\n\n"
    "Final = Site + Unit (capped 1-10)", 11)

row += 2

#===============================================================================
# UNIT MIX & SIZE (5% weight)
#===============================================================================
row = add_section_header(row, 'UNIT MIX & SIZE (5% weight)')
row = add_column_headers(row)

um_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Total Units', xlookup('Number of Units')
ws[f'C{row}'] = xlookup('UM: Total Units')
ws[f'E{row}'] = 'Size Scale'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Avg Unit Size (SF)', xlookup('Avg Unit Size (SF)')
ws[f'C{row}'] = xlookup('UM: Avg SF')
ws[f'E{row}'], ws[f'F{row}'] = '>=1000 SF', 10
row += 1

ws[f'A{row}'], ws[f'B{row}'] = '2BR Units', xlookup('Subject Units - 2BR')
ws[f'E{row}'], ws[f'F{row}'] = '900-1000', 8
row += 1

ws[f'A{row}'], ws[f'B{row}'] = '3BR Units', xlookup('Subject Units - 3BR')
ws[f'E{row}'], ws[f'F{row}'] = '800-900', 6
row += 1

ws[f'A{row}'] = '2-3BR Units'
ws[f'B{row}'] = f'=IF(OR(B{um_start+2}="",B{um_start+3}=""),"",B{um_start+2}+B{um_start+3})'
ws[f'E{row}'], ws[f'F{row}'] = '700-800', 4
row += 1

ws[f'A{row}'] = '2-3BR %'
ws[f'B{row}'] = f'=IF(OR(B{row-1}="",B{um_start}=""),"",ROUND(B{row-1}/B{um_start}*100,1))'
ws[f'C{row}'] = xlookup('UM: 2-3BR %')
ws[f'E{row}'], ws[f'F{row}'] = '<700', 2
row += 1

ws[f'A{row}'] = 'Size Score'
ws[f'B{row}'] = f'=IF(B{um_start+1}="","",IF(B{um_start+1}>=1000,10,IF(B{um_start+1}>=900,8,IF(B{um_start+1}>=800,6,IF(B{um_start+1}>=700,4,2)))))'
ws[f'C{row}'] = xlookup('UM: Size Score')
ws[f'E{row}'] = 'Mix Scale'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'] = 'Mix Score'
ws[f'B{row}'] = f'=IF(B{row-2}="","",IF(B{row-2}>=70,10,IF(B{row-2}>=60,8,IF(B{row-2}>=50,6,IF(B{row-2}>=40,4,3)))))'
ws[f'C{row}'] = xlookup('UM: Mix Score')
ws[f'E{row}'], ws[f'F{row}'] = '>=70% 2-3BR', 10
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(AND(B{row-2}="",B{row-1}=""),"",ROUND(AVERAGE(B{row-2},B{row-1}),0))'
ws[f'B{row}'].fill = final_fill
ws[f'C{row}'] = xlookup('UM: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '60-70%', 8
final_score_rows['UM'] = row
row += 1

ws[f'E{row}'], ws[f'F{row}'] = '50-60%', 6
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '40-50%', 4
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '<40%', 3
row += 1

add_methodology(um_start,
    "LOGIC: Family-friendly units for workforce housing.\n\n"
    "Two sub-scores averaged:\n\n"
    "1. Size Score (avg unit SF)\n"
    "   - Larger units = better for families\n"
    "   - 1000+ SF ideal\n\n"
    "2. Mix Score (% 2-3 bedrooms)\n"
    "   - More 2-3BR = more families\n"
    "   - 70%+ ideal\n\n"
    "Final = Average of both scores", 12)

row += 2

#===============================================================================
# LOSS-TO-LEASE (10% weight)
#===============================================================================
row = add_section_header(row, 'LOSS-TO-LEASE & NOI UPSIDE (10% weight)')
row = add_column_headers(row)

ltl_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'Subject Rent (Avg)', xlookup('Subject Current Rent (Avg)')
ws[f'C{row}'] = xlookup('LTL: Subject Rent')
ws[f'E{row}'] = 'LTL %'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Comp Avg Rent', xlookup('Avg Comp Rent/Unit')
ws[f'C{row}'] = xlookup('LTL: Comp Avg Rent')
ws[f'E{row}'], ws[f'F{row}'] = '<=-20%', 10
row += 1

ws[f'A{row}'], ws[f'B{row}'] = 'Submarket Avg Rent', xlookup('Submarket Avg Rent')
ws[f'C{row}'] = xlookup('LTL: Submarket Rent')
ws[f'E{row}'], ws[f'F{row}'] = '-15 to -20%', 9
row += 1

ws[f'A{row}'] = 'LTL vs Comps %'
ws[f'B{row}'] = f'=IF(OR(B{ltl_start}="",B{ltl_start+1}="",B{ltl_start+1}=0),"N/A",ROUND((B{ltl_start}-B{ltl_start+1})/B{ltl_start+1}*100,1))'
ws[f'C{row}'] = xlookup('LTL: vs Comps %')
ws[f'E{row}'], ws[f'F{row}'] = '-10 to -15%', 8
row += 1

ws[f'A{row}'] = 'LTL vs Submarket %'
ws[f'B{row}'] = f'=IF(OR(B{ltl_start}="",B{ltl_start+2}="",B{ltl_start+2}=0),"N/A",ROUND((B{ltl_start}-B{ltl_start+2})/B{ltl_start+2}*100,1))'
ws[f'C{row}'] = xlookup('LTL: vs Submarket %')
ws[f'E{row}'], ws[f'F{row}'] = '-5 to -10%', 7
row += 1

ws[f'A{row}'] = 'Blended LTL (60/40)'
ws[f'B{row}'] = f'=IF(AND(ISNUMBER(B{row-2}),ISNUMBER(B{row-1})),ROUND(B{row-2}*0.6+B{row-1}*0.4,1),IF(ISNUMBER(B{row-2}),B{row-2},IF(ISNUMBER(B{row-1}),B{row-1},"N/A")))'
ws[f'C{row}'] = xlookup('LTL: Blended %')
ws[f'E{row}'], ws[f'F{row}'] = '-2.5 to -5%', 6
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
ws[f'B{row}'] = f'=IF(NOT(ISNUMBER(B{row-1})),"",IF(B{row-1}<=-20,10,IF(B{row-1}<=-15,9,IF(B{row-1}<=-10,8,IF(B{row-1}<=-5,7,IF(B{row-1}<=-2.5,6,IF(B{row-1}<=2.5,5,IF(B{row-1}<=5,4,IF(B{row-1}<=10,3,IF(B{row-1}<=15,2,IF(B{row-1}<=20,1,0)))))))))))'
ws[f'B{row}'].fill = final_fill
ws[f'C{row}'] = xlookup('LTL: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = '+/- 2.5%', 5
final_score_rows['LTL'] = row
row += 1

ws[f'E{row}'], ws[f'F{row}'] = '+2.5 to +5%', 4
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '+5 to +10%', 3
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '+10 to +15%', 2
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '+15 to +20%', 1
row += 1
ws[f'E{row}'], ws[f'F{row}'] = '>+20%', 0
row += 1

add_methodology(ltl_start,
    "LOGIC: Is there room to raise rents?\n\n"
    "LTL% = (Subject - Market) / Market\n"
    "- Negative = below market (GOOD - upside)\n"
    "- Positive = above market (BAD - limited upside)\n\n"
    "Blended = 60% vs Comps + 40% vs Submarket\n"
    "- Comps weighted higher (direct competitors)\n\n"
    "Example: -10% LTL means rents 10% below market\n"
    "= significant upside potential", 12)

row += 2

#===============================================================================
# BUSINESS-FRIENDLY ENVIRONMENT (3% weight)
#===============================================================================
row = add_section_header(row, 'BUSINESS-FRIENDLY ENVIRONMENT (3% weight)')
row = add_column_headers(row)

bf_start = row
ws[f'A{row}'], ws[f'B{row}'] = 'State', xlookup('State')
ws[f'C{row}'] = xlookup('BF: State')
ws[f'E{row}'] = 'States'
ws[f'E{row}'].font = bold_font
row += 1

ws[f'A{row}'] = 'FINAL SCORE'
ws[f'A{row}'].font = bold_font
state_lookup = f'=LET(st,UPPER(B{bf_start}),IF(OR(st="TX",st="FL",st="TN",st="AZ"),10,IF(OR(st="GA",st="NC",st="SC",st="NV",st="IN"),9,IF(OR(st="KS",st="MO",st="OH",st="UT",st="OK",st="AL"),8,IF(OR(st="CO",st="ID",st="KY",st="AR",st="NE",st="LA",st="MS"),7,IF(OR(st="PA",st="MI",st="WI",st="VA",st="IA",st="MT",st="WY",st="SD",st="ND"),6,IF(OR(st="IL",st="MN",st="NM",st="WV",st="AK"),5,IF(OR(st="WA",st="MD",st="NH",st="DE"),4,IF(OR(st="MA",st="NJ",st="CT",st="HI",st="ME"),3,IF(OR(st="NY",st="VT",st="RI"),2,IF(OR(st="CA",st="OR",st="DC"),1,"")))))))))))'
ws[f'B{row}'] = state_lookup
ws[f'B{row}'].fill = final_fill
ws[f'C{row}'] = xlookup('BF: Final Score')
ws[f'E{row}'], ws[f'F{row}'] = 'TX,FL,TN,AZ', 10
final_score_rows['BF'] = row
row += 1

ws[f'E{row}'], ws[f'F{row}'] = 'GA,NC,SC,NV,IN', 9
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'KS,MO,OH,UT,OK,AL', 8
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'CO,ID,KY,AR,NE...', 7
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'PA,MI,WI,VA,IA...', 6
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'IL,MN,NM,WV,AK', 5
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'WA,MD,NH,DE', 4
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'MA,NJ,CT,HI,ME', 3
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'NY,VT,RI', 2
row += 1
ws[f'E{row}'], ws[f'F{row}'] = 'CA,OR,DC', 1
row += 1

add_methodology(bf_start,
    "LOGIC: How landlord-friendly is the state?\n\n"
    "Factors considered:\n"
    "- Rent control laws (none = better)\n"
    "- Eviction process ease\n"
    "- Tenant protection laws\n"
    "- Business tax environment\n\n"
    "High scores (TX, FL, TN, AZ):\n"
    "- No rent control, easy evictions\n\n"
    "Low scores (CA, OR, NY):\n"
    "- Rent control common, strict tenant laws", 11)

row += 2

# Update Stage 2 references
ws2 = wb['Stage 2']
stage2_rows = {
    'SD': 10, 'SO': 13, 'MG': 14, 'LTL': 17,
    'AM': 19, 'UM': 20, 'PR': 28, 'BF': 12
}
for code, stage2_row in stage2_rows.items():
    sr_row = final_score_rows[code]
    ws2[f'F{stage2_row}'] = f"='Scoring Reference'!B{sr_row}"

# Save
wb.save(r'C:\Users\carso\OneDrive - UCB-O365\Work\RMP\Screener\RMP Screener_PreLinked_v3.xlsx')
print(f'Scoring Reference V2 complete!')
print(f'Final score rows: {final_score_rows}')
