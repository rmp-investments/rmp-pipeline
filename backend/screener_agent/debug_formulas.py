"""Debug script to check formulas in the template Excel file."""
import openpyxl
from openpyxl import load_workbook

template_path = r"C:\Users\carso\OneDrive - UCB-O365\Work\RMP\Screener\RMP Screener_PreLinked_v2.xlsx"

# Load workbook preserving formulas
wb = load_workbook(template_path, data_only=False)

print("=" * 60)
print("FORMULAS IN TEMPLATE")
print("=" * 60)

# Check each sheet for formulas
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    formulas = []

    for row in sheet.iter_rows(min_row=1, max_row=50, max_col=30):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append(f"  {cell.coordinate}: {cell.value[:60]}...")

    if formulas:
        print(f"\n{sheet_name} ({len(formulas)} formulas):")
        for f in formulas[:15]:  # Show first 15
            print(f)
        if len(formulas) > 15:
            print(f"  ... and {len(formulas) - 15} more")

print("\n" + "=" * 60)
print("CELLS WRITTEN BY AGENT (from excel_writer.py)")
print("=" * 60)

# Cells written by write_rent_comps (rows 9-26, many columns)
print("\nRent Comps sheet - Agent writes to:")
rent_comp_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
                 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']
print(f"  Columns: {', '.join(rent_comp_cols)}")
print(f"  Rows: 9-26")

# Check if any of those cells have formulas
if 'Rent Comps' in wb.sheetnames:
    sheet = wb['Rent Comps']
    formula_conflicts = []
    for row in range(9, 27):
        for col in rent_comp_cols:
            cell = sheet[f'{col}{row}']
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_conflicts.append(f"  {col}{row}: {cell.value[:50]}...")

    if formula_conflicts:
        print(f"\n  [!] FORMULA CONFLICTS DETECTED in Rent Comps:")
        for fc in formula_conflicts[:10]:
            print(fc)
    else:
        print("\n  No formula conflicts in data rows 9-26")

# Also check if there are formulas anywhere else in Rent Comps
print("\n  All formulas in Rent Comps sheet:")
for row in range(1, 30):
    for col in rent_comp_cols:
        cell = sheet[f'{col}{row}']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"    {col}{row}: {cell.value[:50]}")

# Data Inputs sheet
print("\nData Inputs sheet - Agent writes to:")
print("  Columns B, C, D, E")
print("  Rows 1-100 (clears and recreates)")
