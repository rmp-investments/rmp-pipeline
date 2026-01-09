"""Compare template formulas to output file to find what changed."""
import openpyxl
from openpyxl import load_workbook

template_path = r"C:\Users\carso\OneDrive - UCB-O365\Work\RMP\Screener\RMP Screener_PreLinked_v2.xlsx"
output_path = r"C:\Users\carso\OneDrive - UCB-O365\Work\RMP\Screener\Properties\Fieldstone\RMP Screener_Fieldstone Apartments.xlsx"

# Load both workbooks
template = load_workbook(template_path, data_only=False)
try:
    output = load_workbook(output_path, data_only=False)
except FileNotFoundError:
    print("Output file not found. Running agent first...")
    exit(1)

print("=" * 60)
print("COMPARING TEMPLATE TO OUTPUT")
print("=" * 60)

# Check each sheet
for sheet_name in template.sheetnames:
    if sheet_name not in output.sheetnames:
        print(f"\n{sheet_name}: MISSING from output!")
        continue

    template_sheet = template[sheet_name]
    output_sheet = output[sheet_name]

    differences = []

    # Check all cells for formula differences
    for row in range(1, 50):
        for col in range(1, 30):
            t_cell = template_sheet.cell(row=row, column=col)
            o_cell = output_sheet.cell(row=row, column=col)

            t_val = t_cell.value
            o_val = o_cell.value

            # Check if template had a formula
            if t_val and isinstance(t_val, str) and t_val.startswith('='):
                if o_val != t_val:
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    differences.append((cell_ref, t_val, o_val))

    if differences:
        print(f"\n{sheet_name}: {len(differences)} formula changes detected")
        for cell, template_val, output_val in differences[:10]:
            print(f"  {cell}:")
            print(f"    Template: {template_val[:50] if template_val else 'None'}")
            print(f"    Output:   {str(output_val)[:50] if output_val else 'None'}")
    else:
        print(f"\n{sheet_name}: All formulas preserved")

template.close()
output.close()
