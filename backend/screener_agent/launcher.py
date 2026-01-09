"""
RMP Screener Launcher - Simple menu for non-technical users
Double-click RUN_SCREENER.bat in the Screener folder to use
"""

import os
import sys
import json

# Rich console for beautiful output
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.text import Text
from rich.prompt import Prompt, IntPrompt
from rich import box
from rich.align import Align

# Add modules directory to path
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))
from gis_utils import try_get_parcel_data
from data_validator import DataValidator, load_api_key_from_config, save_api_key_to_config

console = Console()

def clear_screen():
    """Clear screen - works on Windows."""
    import os
    os.system('cls' if os.name == 'nt' else 'clear')

# ASCII Art Banner - Using dark_orange for true orange color
BANNER = """[bold dark_orange]
  ██████╗ ███╗   ███╗██████╗     ███████╗ ██████╗██████╗ ███████╗███████╗███╗   ██╗███████╗██████╗
  ██╔══██╗████╗ ████║██╔══██╗    ██╔════╝██╔════╝██╔══██╗██╔════╝██╔════╝████╗  ██║██╔════╝██╔══██╗
  ██████╔╝██╔████╔██║██████╔╝    ███████╗██║     ██████╔╝█████╗  █████╗  ██╔██╗ ██║█████╗  ██████╔╝
  ██╔══██╗██║╚██╔╝██║██╔═══╝     ╚════██║██║     ██╔══██╗██╔══╝  ██╔══╝  ██║╚██╗██║██╔══╝  ██╔══██╗
  ██║  ██║██║ ╚═╝ ██║██║         ███████║╚██████╗██║  ██║███████╗███████╗██║ ╚████║███████╗██║  ██║
  ╚═╝  ╚═╝╚═╝     ╚═╝╚═╝         ╚══════╝ ╚═════╝╚═╝  ╚═╝╚══════╝╚══════╝╚═╝  ╚═══╝╚══════╝╚═╝  ╚═╝
[/bold dark_orange][dim rgb(205,102,0)]                         ══════════════════════════════════════════════════════════
                                     Property Analysis Automation Tool  v2.0
                         ══════════════════════════════════════════════════════════[/dim rgb(205,102,0)]
"""

# Base directories - all paths are relative to the screener_agent folder
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))  # screener_agent folder
BASE_DIR = os.path.dirname(SCRIPT_DIR)  # Screener folder (parent of screener_agent)

def resolve_path(path):
    """Resolve a path - if relative, resolve from SCRIPT_DIR."""
    if path is None:
        return None
    if os.path.isabs(path):
        return path
    return os.path.normpath(os.path.join(SCRIPT_DIR, path))

# Load settings from settings.json
SETTINGS_FILE = os.path.join(SCRIPT_DIR, 'settings.json')
if os.path.exists(SETTINGS_FILE):
    with open(SETTINGS_FILE, 'r') as f:
        _settings = json.load(f)
    PROPERTIES_DIR = resolve_path(_settings.get('properties_dir', '../Properties'))
    TEMPLATE_EXCEL = resolve_path(_settings.get('template_excel', '../RMP Screener_PreLinked_v3.xlsx'))
    CONFIGS_DIR = _settings.get('configs_dir') or os.path.join(SCRIPT_DIR, 'configs')
else:
    PROPERTIES_DIR = os.path.join(BASE_DIR, 'Properties')
    TEMPLATE_EXCEL = os.path.join(BASE_DIR, 'RMP Screener_PreLinked_v3.xlsx')
    CONFIGS_DIR = os.path.join(SCRIPT_DIR, 'configs')


def is_file_open(filepath):
    """Check if a file is currently open by another process (Windows)."""
    if not os.path.exists(filepath):
        return False
    try:
        # Try to open file with exclusive access
        with open(filepath, 'r+b'):
            pass
        return False
    except (IOError, PermissionError):
        return True


def find_excel_file(config, output_folder):
    """Find the Excel file for a property."""
    import glob
    paths = config.get('paths', {})
    possible_paths = [
        paths.get('output_file', ''),
        paths.get('screener_file', ''),  # Also check screener_file
        os.path.join(output_folder, f"RMP Screener_{config['property_name']}.xlsx"),
    ]
    # Search for any screener xlsx in the folder
    screener_files = glob.glob(os.path.join(output_folder, "RMP Screener*.xlsx"))
    possible_paths.extend(screener_files)

    for path in possible_paths:
        if path and os.path.exists(path):
            return path
    return None


def update_excel_parcel(paths, output_folder, config, screenshot_path, console):
    """Update Excel file with new parcel screenshot, replacing any existing parcel image."""
    import glob

    excel_path = None
    possible_paths = [
        paths.get('output_file', ''),
        paths.get('screener_file', ''),  # Also check screener_file
        os.path.join(output_folder, f"RMP Screener_{config['property_name']}.xlsx"),
    ]
    # Also search for any screener xlsx in the folder
    screener_files = glob.glob(os.path.join(output_folder, "RMP Screener*.xlsx"))
    possible_paths.extend(screener_files)

    for path in possible_paths:
        if path and os.path.exists(path):
            excel_path = path
            break

    if excel_path:
        console.print(f"[dim]Updating: {excel_path}[/dim]")
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as XLImage

            wb = load_workbook(excel_path)
            if 'Screener Cover' in wb.sheetnames:
                ws = wb['Screener Cover']

                # Remove existing images at or near E39 (parcel map location)
                # openpyxl stores images in ws._images list
                images_to_keep = []
                for img in ws._images:
                    # Check if image is anchored at E39 area (parcel map)
                    anchor = img.anchor
                    if hasattr(anchor, '_from'):
                        # TwoCellAnchor - check row/col
                        col = anchor._from.col
                        row = anchor._from.row
                        # E39 = col 4 (0-indexed), row 38 (0-indexed)
                        # Keep images NOT in the parcel area (E39 region)
                        if not (col == 4 and 35 <= row <= 45):
                            images_to_keep.append(img)
                    elif hasattr(anchor, 'col') and hasattr(anchor, 'row'):
                        col = anchor.col
                        row = anchor.row
                        if not (col == 4 and 35 <= row <= 45):
                            images_to_keep.append(img)
                    else:
                        # String anchor like 'E39' - parse it
                        anchor_str = str(anchor) if anchor else ''
                        if anchor_str.startswith('E') and any(c.isdigit() for c in anchor_str):
                            try:
                                row_num = int(''.join(c for c in anchor_str if c.isdigit()))
                                if not (35 <= row_num <= 45):
                                    images_to_keep.append(img)
                            except:
                                images_to_keep.append(img)
                        else:
                            images_to_keep.append(img)

                # Replace the images list
                removed_count = len(ws._images) - len(images_to_keep)
                ws._images = images_to_keep
                if removed_count > 0:
                    console.print(f"[dim]Removed {removed_count} old parcel image(s)[/dim]")

                # Add new parcel image
                xl_img = XLImage(screenshot_path)
                xl_img.anchor = 'E39'
                xl_img.width = 4.0 * 96  # 4 inches
                xl_img.height = int(xl_img.width * (700/800))
                ws.add_image(xl_img)

                wb.save(excel_path)
                console.print(f"[green]Excel updated![/green]")
            else:
                console.print("[yellow]Screener Cover sheet not found[/yellow]")
            wb.close()
        except Exception as e:
            console.print(f"[red]Could not update Excel: {e}[/red]")
            console.print("[dim]Make sure Excel file is closed.[/dim]")
    else:
        console.print(f"[yellow]No Excel file found in {output_folder}[/yellow]")


def get_configured_properties():
    """Get list of configured properties."""
    properties = []
    if os.path.exists(CONFIGS_DIR):
        for f in os.listdir(CONFIGS_DIR):
            # Skip template file
            if f.endswith('.json') and f != 'template.json':
                slug = f.replace('.json', '')
                with open(os.path.join(CONFIGS_DIR, f), 'r') as file:
                    config = json.load(file)
                    properties.append({
                        'slug': slug,
                        'name': config.get('property_name', slug)
                    })
    return properties


def show_menu():
    """Display main menu."""
    clear_screen()
    console.print(BANNER)

    # Menu options table
    menu_table = Table(show_header=False, box=box.ROUNDED, border_style="rgb(205,102,0)",
                      padding=(0, 3), expand=False)
    menu_table.add_column("Option", style="bold dark_orange", justify="center", width=4)
    menu_table.add_column("Description", style="white")

    menu_table.add_row("1", "Run screener for existing property")
    menu_table.add_row("2", "Set up a NEW property")
    menu_table.add_row("3", "Fix Property Location")
    menu_table.add_row("4", "Validate Output (uses Claude API)")
    menu_table.add_row("E", "Exit")

    console.print(Panel(menu_table, title="[bold white]Main Menu[/bold white]",
                       border_style="dark_orange", box=box.DOUBLE))
    console.print()


def run_screener():
    """Run screener for an existing property."""
    clear_screen()
    console.print(BANNER)
    properties = get_configured_properties()

    if not properties:
        console.print()
        console.print(Panel(
            "[yellow]No properties configured yet![/yellow]\n\nPlease set up a new property first (option 2).",
            title="[bold yellow]Notice[/bold yellow]",
            border_style="yellow",
            box=box.ROUNDED
        ))
        input("\nPress Enter to continue...")
        return

    # Property selection table
    console.print()
    prop_table = Table(title="[bold]Select a Property[/bold]", box=box.ROUNDED,
                      border_style="rgb(205,102,0)", show_header=True, header_style="bold dark_orange")
    prop_table.add_column("#", style="dark_orange", justify="center", width=4)
    prop_table.add_column("Property Name", style="white")

    for i, prop in enumerate(properties, 1):
        prop_table.add_row(str(i), prop['name'])

    console.print(prop_table)
    console.print()

    choice = Prompt.ask("[dark_orange]Enter number[/dark_orange] [dim](or 'b' to go back)[/dim]")

    if choice.lower() == 'b':
        return

    try:
        idx = int(choice) - 1
        if 0 <= idx < len(properties):
            prop = properties[idx]

            console.print()
            console.print(Panel(
                f"[bold green]Running screener for:[/bold green] [bold yellow]{prop['name']}[/bold yellow]",
                border_style="green",
                box=box.ROUNDED
            ))

            # Import and run agent (v2 parallel version)
            from agent_v2 import ScreenerAgent
            config_path = os.path.join(CONFIGS_DIR, f"{prop['slug']}.json")
            agent = ScreenerAgent(config_path)
            agent.run()

            # Success message
            console.print()
            success_panel = Panel(
                f"[bold green]COMPLETE![/bold green]\n\n"
                f"[white]Check the Excel file in:[/white]\n"
                f"[dark_orange]{PROPERTIES_DIR}\\{prop['name'].replace(' ', '')}\\[/dark_orange]",
                title="[bold white]Success[/bold white]",
                border_style="green",
                box=box.DOUBLE
            )
            console.print(success_panel)
            input("\nPress Enter to continue...")
        else:
            console.print("[red]Invalid selection.[/red]")
    except ValueError:
        console.print("[red]Invalid input.[/red]")
    except Exception as e:
        console.print()
        console.print(Panel(
            f"[red]{str(e)}[/red]",
            title="[bold red]Error[/bold red]",
            border_style="red",
            box=box.ROUNDED
        ))
        input("\nPress Enter to continue...")


def setup_new_property():
    """Set up a new property."""
    clear_screen()
    console.print(BANNER)
    console.print(Panel(
        "[bold]Enter property details below[/bold]\n[dim]Type 'b' at any prompt to go back[/dim]",
        title="[bold white]New Property Setup[/bold white]",
        border_style="dark_orange",
        box=box.DOUBLE
    ))
    console.print()

    # Property name
    name = Prompt.ask("[dark_orange]Property Name[/dark_orange] [dim](e.g., 'Sunset Ridge Apartments')[/dim]")
    if name.lower() == 'b':
        return
    if not name:
        console.print("[red]Property name is required.[/red]")
        input("\nPress Enter to continue...")
        return

    # Address
    console.print()
    console.print("[bold]Property Address:[/bold]")
    address = Prompt.ask("  [dark_orange]Street Address[/dark_orange]")
    city = Prompt.ask("  [dark_orange]City[/dark_orange]")
    state = Prompt.ask("  [dark_orange]State[/dark_orange] [dim](2-letter code)[/dim]").upper()
    zip_code = Prompt.ask("  [dark_orange]ZIP Code[/dark_orange]")

    if not all([address, city, state, zip_code]):
        console.print("\n[red]All address fields are required.[/red]")
        input("\nPress Enter to continue...")
        return

    # Create slug from name
    slug = name.lower().replace(' ', '_').replace('-', '_')
    slug = ''.join(c for c in slug if c.isalnum() or c == '_')

    # Create folder structure
    property_folder = os.path.join(PROPERTIES_DIR, name.replace(' ', ''))
    costar_folder = os.path.join(property_folder, 'CoStar Reports')

    os.makedirs(costar_folder, exist_ok=True)

    # Create config file
    config = {
        'property_name': name,
        'property_slug': slug,
        'property_details': {
            'address': address,
            'city': city,
            'state': state,
            'zip_code': zip_code
        },
        'costar_reports_path': costar_folder,
        'output_path': property_folder
    }

    config_path = os.path.join(CONFIGS_DIR, f'{slug}.json')
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=2)

    # Success message with next steps
    console.print()

    next_steps = Table(show_header=False, box=None, padding=(0, 2))
    next_steps.add_column("Step", style="yellow")
    next_steps.add_column("Action", style="white")

    next_steps.add_row("1.", f"Add CoStar PDF reports to:\n   [dark_orange]{costar_folder}[/dark_orange]")
    next_steps.add_row("", "")
    next_steps.add_row("", "[dim]Required PDFs:[/dim]")
    next_steps.add_row("", "  [white]- Demographic Report.pdf[/white]")
    next_steps.add_row("", "  [white]- Property Report.pdf[/white]")
    next_steps.add_row("", "  [white]- Rent Comp Report.pdf[/white]")
    next_steps.add_row("", "  [white]- Asset Market Report.pdf[/white]")
    next_steps.add_row("", "")
    next_steps.add_row("2.", "Run this launcher again and select option 1")

    success_content = Text()
    success_content.append("PROPERTY CREATED!\n\n", style="bold green")
    success_content.append(f"Folder: ", style="white")
    success_content.append(f"{property_folder}\n\n", style="dark_orange")

    console.print(Panel(
        success_content,
        title="[bold white]Success[/bold white]",
        border_style="green",
        box=box.DOUBLE
    ))

    console.print()
    console.print(Panel(
        next_steps,
        title="[bold white]Next Steps[/bold white]",
        border_style="yellow",
        box=box.ROUNDED
    ))

    input("\nPress Enter to continue...")


def fix_parcel_map():
    """Fix property location by letting user manually adjust the map center."""
    clear_screen()
    console.print(BANNER)
    properties = get_configured_properties()

    if not properties:
        console.print()
        console.print(Panel(
            "[yellow]No properties configured yet![/yellow]\n\nPlease set up a new property first.",
            title="[bold yellow]Notice[/bold yellow]",
            border_style="yellow",
            box=box.ROUNDED
        ))
        input("\nPress Enter to continue...")
        return

    # Property selection table
    console.print()
    prop_table = Table(title="[bold]Select Property to Fix Location[/bold]", box=box.ROUNDED,
                      border_style="rgb(205,102,0)", show_header=True, header_style="bold dark_orange")
    prop_table.add_column("#", style="dark_orange", justify="center", width=4)
    prop_table.add_column("Property Name", style="white")

    for i, prop in enumerate(properties, 1):
        prop_table.add_row(str(i), prop['name'])

    console.print(prop_table)
    console.print()

    choice = Prompt.ask("[dark_orange]Enter number[/dark_orange] [dim](or 'b' to go back)[/dim]")

    if choice.lower() == 'b':
        return

    try:
        idx = int(choice) - 1
        if 0 <= idx < len(properties):
            prop = properties[idx]
            config_path = os.path.join(CONFIGS_DIR, f"{prop['slug']}.json")

            # Load config
            with open(config_path, 'r') as f:
                config = json.load(f)

            # Determine output folder for Excel check
            paths = config.get('paths', {})
            output_folder = paths.get('output_path')
            if not output_folder:
                output_file = paths.get('output_file', '')
                if output_file:
                    output_folder = os.path.dirname(output_file)
            if not output_folder:
                costar_dir = paths.get('costar_reports_dir', '')
                if costar_dir:
                    output_folder = os.path.dirname(costar_dir)
            if not output_folder:
                output_folder = os.path.join(PROPERTIES_DIR, prop['name'].replace(' ', ''))

            # Check if Excel file is open
            excel_path = find_excel_file(config, output_folder)
            if excel_path and is_file_open(excel_path):
                console.print()
                console.print(Panel(
                    f"[bold yellow]⚠ WARNING: Excel file is currently open![/bold yellow]\n\n"
                    f"[white]File: {os.path.basename(excel_path)}[/white]\n\n"
                    "[white]Please close the Excel file before continuing.[/white]\n"
                    "[white]The screener needs to update the parcel image in the file.[/white]",
                    title="[bold yellow]Close Excel[/bold yellow]",
                    border_style="yellow",
                    box=box.DOUBLE
                ))
                input("\nPress Enter after closing Excel to continue...")

                # Check again
                if is_file_open(excel_path):
                    console.print("[red]Excel file still appears to be open. Continuing anyway...[/red]")
                    console.print("[dim]The Excel update may fail - you can manually replace the parcel image later.[/dim]")

            console.print()
            console.print(Panel(
                f"[bold cyan]Fixing property location for:[/bold cyan] [bold yellow]{prop['name']}[/bold yellow]\n\n"
                "[white]A browser window will open with a zoomed-out satellite view.[/white]\n"
                "[white]Pan and zoom to center on the property parcel.[/white]\n"
                "[white]Then come back here and press Enter to capture.[/white]",
                border_style="cyan",
                box=box.ROUNDED
            ))

            input("\nPress Enter to open the map...")

            # Import map generator
            from modules.map_generator import MapGenerator

            # Get coordinates (from config or geocode)
            prop_details = config.get('property_details', {})
            lat = prop_details.get('parcel_lat') or prop_details.get('latitude')
            lon = prop_details.get('parcel_lon') or prop_details.get('longitude')

            if not lat or not lon:
                # Geocode the address
                import requests
                address = f"{prop_details['address']}, {prop_details['city']}, {prop_details['state']} {prop_details['zip_code']}"
                try:
                    url = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress"
                    params = {'address': address, 'benchmark': 'Public_AR_Current', 'format': 'json'}
                    resp = requests.get(url, params=params, timeout=15)
                    data = resp.json()
                    matches = data.get('result', {}).get('addressMatches', [])
                    if matches:
                        coords = matches[0]['coordinates']
                        lat, lon = coords['y'], coords['x']
                except Exception as e:
                    console.print(f"[red]Could not geocode address: {e}[/red]")
                    input("\nPress Enter to continue...")
                    return

            if not lat or not lon:
                console.print("[red]Could not determine property coordinates.[/red]")
                input("\nPress Enter to continue...")
                return

            # PRIORITY 1: Try to auto-detect parcel from county GIS
            console.print("[dark_orange]Attempting to auto-detect parcel from county GIS...[/dark_orange]")
            street_address = prop_details.get('address', '')
            parcel_data = try_get_parcel_data(lat, lon, property_address=street_address, console=console)
            use_auto = 'n'  # Default to manual

            if parcel_data:
                centroid_lat, centroid_lon = parcel_data['centroid']
                parcel_polygon = parcel_data['polygon']
                console.print(f"[green]Found parcel: {centroid_lat:.6f}, {centroid_lon:.6f} ({len(parcel_polygon)} vertices)[/green]")

                # Ask user if they want to use this or manually adjust
                console.print()
                use_auto = input("Use auto-detected parcel? (y/n, default=y): ").strip().lower()
                if use_auto == '':
                    use_auto = 'y'

                if use_auto != 'n':
                    # Save coordinates and polygon, skip manual adjustment
                    config['property_details']['parcel_lat'] = centroid_lat
                    config['property_details']['parcel_lon'] = centroid_lon
                    config['property_details']['parcel_zoom'] = 18
                    config['property_details']['parcel_polygon'] = parcel_polygon

                    with open(config_path, 'w') as f:
                        json.dump(config, f, indent=2)

                    console.print("[green]Parcel data saved to config![/green]")

                    # Generate the parcel screenshot automatically
                    console.print("[dark_orange]Generating parcel screenshot with boundary...[/dark_orange]")

                    # Set up output paths
                    paths = config.get('paths', {})
                    output_folder = paths.get('output_path')
                    if not output_folder:
                        output_file = paths.get('output_file', '')
                        if output_file:
                            output_folder = os.path.dirname(output_file)
                    if not output_folder:
                        costar_dir = paths.get('costar_reports_dir', '')
                        if costar_dir:
                            output_folder = os.path.dirname(costar_dir)
                    if not output_folder:
                        output_folder = os.path.join(PROPERTIES_DIR, prop['name'].replace(' ', ''))
                    maps_dir = os.path.join(output_folder, 'Maps')
                    os.makedirs(maps_dir, exist_ok=True)

                    # Generate parcel map with polygon outline using MapGenerator
                    # Let zoom auto-calculate from polygon size
                    from modules.map_generator import MapGenerator
                    generator = MapGenerator(centroid_lat, centroid_lon, config['property_name'], maps_dir)
                    parcel_path = generator.create_parcel_satellite(zoom=None, parcel_polygon=parcel_polygon)

                    if parcel_path:
                        console.print(f"[green]Parcel screenshot saved with boundary![/green]")

                        # Update Excel
                        update_excel_parcel(paths, output_folder, config, parcel_path, console)

                        console.print()
                        console.print(Panel(
                            "[bold green]Property location fixed![/bold green]\n\n"
                            f"[white]New coordinates: {centroid_lat:.6f}, {centroid_lon:.6f}[/white]\n"
                            f"[white]Parcel boundary: {len(parcel_polygon)} vertices[/white]\n"
                            "[white]Future runs will use these coordinates.[/white]",
                            border_style="green",
                            box=box.DOUBLE
                        ))
                    else:
                        console.print("[yellow]Could not generate screenshot, falling back to manual...[/yellow]")
                        parcel_data = None  # Fall through to manual

                    if parcel_data:
                        input("\nPress Enter to continue...")
                        return

            # FALLBACK: Manual adjustment
            if not parcel_data or use_auto == 'n':
                console.print("[dim]Falling back to manual adjustment...[/dim]")

            # Set up output paths - use same logic as agent
            paths = config.get('paths', {})
            output_folder = paths.get('output_path')
            if not output_folder:
                output_file = paths.get('output_file', '')
                if output_file:
                    output_folder = os.path.dirname(output_file)
            if not output_folder:
                costar_dir = paths.get('costar_reports_dir', '')
                if costar_dir:
                    output_folder = os.path.dirname(costar_dir)
            if not output_folder:
                output_folder = os.path.join(PROPERTIES_DIR, prop['name'].replace(' ', ''))
            maps_dir = os.path.join(output_folder, 'Maps')
            os.makedirs(maps_dir, exist_ok=True)

            # Create zoomed-out interactive map for adjustment
            generator = MapGenerator(lat, lon, config['property_name'], maps_dir)

            console.print("[dark_orange]Creating zoomed-out map for adjustment...[/dark_orange]")
            interactive_html = generator.create_parcel_for_adjustment(zoom=16)  # Zoomed out more

            if not interactive_html:
                console.print("[red]Failed to create adjustment map.[/red]")
                input("\nPress Enter to continue...")
                return

            # Open in browser (not headless) using Selenium
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            import time

            chrome_options = Options()
            # NOT headless - user can see and interact
            # Use a large window size instead of maximized (avoids resize issues)
            chrome_options.add_argument('--window-size=1200,900')

            console.print("[green]Opening browser...[/green]")
            driver = webdriver.Chrome(options=chrome_options)
            driver.get(f'file:///{interactive_html}')

            console.print()
            console.print(Panel(
                "[bold green]Browser is open![/bold green]\n\n"
                "[white]1. Pan and zoom the map to center the parcel on the crosshair[/white]\n"
                "[white]2. Click the green DONE button when ready[/white]\n"
                "[white]3. Keep browser open and press Enter here to capture[/white]",
                border_style="green",
                box=box.ROUNDED
            ))

            input("\nPress Enter after clicking DONE in browser...")

            # Get current map center from browser
            try:
                center_lat = driver.execute_script("return map.getCenter().lat;")
                center_lon = driver.execute_script("return map.getCenter().lng;")
                current_zoom = driver.execute_script("return map.getZoom();")
                console.print(f"[dim]Captured view: {center_lat:.6f}, {center_lon:.6f} (zoom {current_zoom})[/dim]")
            except:
                center_lat, center_lon = lat, lon
                current_zoom = 18

            # Capture screenshot
            console.print("[dark_orange]Capturing screenshot...[/dark_orange]")

            # Set window size for consistent capture (restore from maximized if needed)
            try:
                driver.set_window_size(800, 700)
            except:
                # Window might be maximized - restore it first
                driver.set_window_position(100, 100)  # Forces out of maximized state
                driver.set_window_size(800, 700)
            time.sleep(0.5)

            # IMPORTANT: Re-center map to captured coordinates (resize can change the view)
            try:
                driver.execute_script(f"map.setView([{center_lat}, {center_lon}], {current_zoom});")
            except:
                pass  # Map variable may not be accessible - proceed anyway
            time.sleep(1)  # Let map settle

            # Hide all overlay elements before screenshot
            hide_js = """
                // Hide specific overlay elements by ID
                var instructions = document.getElementById('instructions');
                var doneBtn = document.getElementById('doneBtn');
                if (instructions) instructions.style.visibility = 'hidden';
                if (doneBtn) doneBtn.style.visibility = 'hidden';

                // Hide by class (crosshair, center-dot, crop overlay, etc)
                document.querySelectorAll('.instructions, .done-btn, .crosshair, .crosshair-h, .crosshair-v, .center-dot, .crop-overlay, .crop-label').forEach(function(el) {
                    el.style.visibility = 'hidden';
                });

                // Also try removing them entirely
                document.querySelectorAll('button').forEach(function(el) {
                    el.remove();
                });
                document.querySelectorAll('.instructions, .crosshair, .crop-overlay, .crop-label').forEach(function(el) {
                    el.remove();
                });
            """
            driver.execute_script(hide_js)
            time.sleep(0.3)

            screenshot_path = os.path.join(maps_dir, f"{config['property_name']}_parcel.png")
            driver.save_screenshot(screenshot_path)
            driver.quit()

            console.print(f"[green]Screenshot saved:[/green] {screenshot_path}")

            # Save coordinates to config
            config['property_details']['parcel_lat'] = center_lat
            config['property_details']['parcel_lon'] = center_lon
            config['property_details']['parcel_zoom'] = current_zoom

            with open(config_path, 'w') as f:
                json.dump(config, f, indent=2)

            console.print("[green]Coordinates saved to config[/green]")

            # Update Excel file using the shared function
            console.print("[dark_orange]Updating Excel file...[/dark_orange]")
            update_excel_parcel(paths, output_folder, config, screenshot_path, console)

            console.print()
            console.print(Panel(
                "[bold green]Property location fixed![/bold green]\n\n"
                f"[white]New coordinates saved: {center_lat:.6f}, {center_lon:.6f}[/white]\n"
                "[white]Future runs will use these coordinates automatically.[/white]",
                border_style="green",
                box=box.DOUBLE
            ))

            input("\nPress Enter to continue...")
        else:
            console.print("[red]Invalid selection.[/red]")
    except ValueError:
        console.print("[red]Invalid input.[/red]")
    except Exception as e:
        console.print()
        console.print(Panel(
            f"[red]{str(e)}[/red]",
            title="[bold red]Error[/bold red]",
            border_style="red",
            box=box.ROUNDED
        ))
        import traceback
        traceback.print_exc()
        input("\nPress Enter to continue...")


def validate_output():
    """Validate screener output using Claude API."""
    clear_screen()
    console.print(BANNER)
    properties = get_configured_properties()

    if not properties:
        console.print()
        console.print(Panel(
            "[yellow]No properties configured yet![/yellow]\n\nPlease run the screener first.",
            title="[bold yellow]Notice[/bold yellow]",
            border_style="yellow",
            box=box.ROUNDED
        ))
        input("\nPress Enter to continue...")
        return

    # Property selection table
    console.print()
    prop_table = Table(title="[bold]Select Property to Validate[/bold]", box=box.ROUNDED,
                      border_style="rgb(205,102,0)", show_header=True, header_style="bold dark_orange")
    prop_table.add_column("#", style="dark_orange", justify="center", width=4)
    prop_table.add_column("Property Name", style="white")

    for i, prop in enumerate(properties, 1):
        prop_table.add_row(str(i), prop['name'])

    console.print(prop_table)
    console.print()

    choice = Prompt.ask("[dark_orange]Enter number[/dark_orange] [dim](or 'b' to go back)[/dim]")

    if choice.lower() == 'b':
        return

    try:
        idx = int(choice) - 1
        if 0 <= idx < len(properties):
            prop = properties[idx]
            config_path = os.path.join(CONFIGS_DIR, f"{prop['slug']}.json")

            # Load config to find Excel file
            with open(config_path, 'r') as f:
                config = json.load(f)

            # Find output folder
            paths = config.get('paths', {})
            output_folder = paths.get('output_path')
            if not output_folder:
                output_folder = os.path.join(PROPERTIES_DIR, prop['name'].replace(' ', ''))

            # Find Excel file
            excel_path = find_excel_file(config, output_folder)

            if not excel_path:
                console.print(Panel(
                    "[yellow]No Excel output found for this property.[/yellow]\n\nRun the screener first.",
                    title="[bold yellow]Notice[/bold yellow]",
                    border_style="yellow",
                    box=box.ROUNDED
                ))
                input("\nPress Enter to continue...")
                return

            # Check if file is open
            if is_file_open(excel_path):
                console.print(Panel(
                    f"[bold yellow]⚠ Excel file is open![/bold yellow]\n\n"
                    f"[white]Please close: {os.path.basename(excel_path)}[/white]",
                    title="[bold yellow]Close Excel[/bold yellow]",
                    border_style="yellow",
                    box=box.DOUBLE
                ))
                input("\nPress Enter after closing Excel...")
                if is_file_open(excel_path):
                    console.print("[red]File still open. Cannot proceed.[/red]")
                    input("\nPress Enter to continue...")
                    return

            # Get API key
            api_key = load_api_key_from_config()
            if api_key:
                masked = api_key[:8] + '...' + api_key[-4:]
                console.print(f"\n[dim]Using saved API key: {masked}[/dim]")
            else:
                console.print("\n[yellow]No API key found.[/yellow]")
                console.print("Enter your Claude API key (starts with 'sk-ant-'):")
                api_key = Prompt.ask("[dark_orange]API Key[/dark_orange]")
                if not api_key:
                    console.print("[red]API key required.[/red]")
                    input("\nPress Enter to continue...")
                    return
                # Save for future use
                save_choice = Prompt.ask("Save key for future use?", choices=["y", "n"], default="y")
                if save_choice == 'y':
                    save_api_key_to_config(api_key)

            console.print()
            console.print(Panel(
                f"[bold cyan]Validating:[/bold cyan] [yellow]{prop['name']}[/yellow]\n\n"
                f"[white]File: {os.path.basename(excel_path)}[/white]\n"
                f"[white]Validates: CoStar PDF sources only[/white]\n"
                f"[dim](Web/scraped sources skipped - can't reliably re-verify)[/dim]",
                border_style="cyan",
                box=box.ROUNDED
            ))

            # Run validation - use Excel's actual parent folder
            console.print()
            actual_folder = os.path.dirname(excel_path)
            validator = DataValidator(excel_path, actual_folder)
            validator.init_claude_client(api_key)
            validator.load_excel()

            # Check if we have cached results
            if validator.validation_cache:
                console.print(f"[dim]Found {len(validator.validation_cache)} previously validated fields[/dim]")
                skip_choice = Prompt.ask(
                    "Skip unchanged fields that were previously OK?",
                    choices=["y", "n"],
                    default="y"
                )
                validator.skip_cached_ok = (skip_choice == 'y')

            validator.run_validation()

            # Summary
            summary = validator.get_summary()
            console.print()

            # Color-code summary
            summary_text = (
                f"[green]OK: {summary.get('OK', 0)}[/green]  "
                f"[red]FAIL: {summary.get('FAIL', 0)}[/red]  "
                f"[yellow]?: {summary.get('?', 0)}[/yellow]  "
                f"[dim]SKIP: {summary.get('SKIP', 0)}[/dim]"
            )

            # Include cache info in summary
            cache_info = ""
            if validator.fields_skipped_from_cache > 0:
                cache_info = f"\n[cyan]Cached: {validator.fields_skipped_from_cache} fields skipped (unchanged OKs)[/cyan]"

            console.print(Panel(
                f"[bold]Validation Complete[/bold]\n\n{summary_text}{cache_info}\n\n"
                f"[dim]API calls: {validator.api_calls_made}, Est. cost: ${validator.estimated_cost:.3f}[/dim]",
                border_style="green" if summary.get('FAIL', 0) == 0 else "yellow",
                box=box.DOUBLE
            ))

            # Show failures
            failures = [r for r in validator.validation_results if r['result'] == 'FAIL']
            if failures:
                console.print()
                console.print("[bold red]Failures:[/bold red]")
                for f in failures:
                    console.print(f"  [red]•[/red] {f['field']}: {f['explanation']}")

            # Save
            save_choice = Prompt.ask("\nSave results to Excel?", choices=["y", "n"], default="y")
            if save_choice == 'y':
                validator.save()
                console.print("[green]Results saved![/green]")

            input("\nPress Enter to continue...")
        else:
            console.print("[red]Invalid selection.[/red]")
    except ValueError:
        console.print("[red]Invalid input.[/red]")
    except Exception as e:
        console.print()
        console.print(Panel(
            f"[red]{str(e)}[/red]",
            title="[bold red]Error[/bold red]",
            border_style="red",
            box=box.ROUNDED
        ))
        import traceback
        traceback.print_exc()
        input("\nPress Enter to continue...")


def main():
    """Main launcher loop."""
    try:
        while True:
            show_menu()
            choice = Prompt.ask("[bold dark_orange]Enter choice[/bold dark_orange]", choices=["1", "2", "3", "4", "e", "E"], show_choices=False)

            if choice == '1':
                run_screener()
            elif choice == '2':
                setup_new_property()
            elif choice == '3':
                fix_parcel_map()
            elif choice == '4':
                validate_output()
            elif choice.lower() == 'e':
                console.print()
                console.print(Panel(
                    "[bold]Goodbye![/bold]\n[dim]Thanks for using RMP Screener[/dim]",
                    border_style="dark_orange",
                    box=box.ROUNDED
                ))
                console.print()
                break
    except KeyboardInterrupt:
        console.print("\n[yellow]Exiting...[/yellow]")


if __name__ == "__main__":
    main()
